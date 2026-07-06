import os
import re
import uuid
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from database.database import get_connection
from database.audit_manager import AuditManager
from database.recipe_status_history_manager import RecipeStatusHistoryManager


def _configured_max_upload_bytes():
    try:
        upload_mb = int(os.getenv("CRS_MAX_UPLOAD_MB", "25"))
    except (TypeError, ValueError):
        upload_mb = 25
    return max(1, upload_mb) * 1024 * 1024


class RecipeExcelImportExportManager:
    """Current CRS recipe Excel import/export for recipes, parameters and phase control."""

    SECOND_STAGE_RECIPE_GROUPS = (
        "CAP_STRIP_SIDE",
        "BT_SIDE",
    )

    @staticmethod
    def _is_second_stage(stage_type):
        return (
            str(stage_type or "").strip().upper().replace(" ", "_")
            in {"SECOND_STAGE", "SECONDSTAGE", "SS"}
        )

    @staticmethod
    def _phase_group_filter_sql(stage_type, column_sql):
        if not RecipeExcelImportExportManager._is_second_stage(stage_type):
            return "", []
        placeholders = ", ".join(
            "?"
            for _ in RecipeExcelImportExportManager.SECOND_STAGE_RECIPE_GROUPS
        )
        return (
            f" AND UPPER(COALESCE({column_sql}, 'MAIN')) IN ({placeholders})",
            list(RecipeExcelImportExportManager.SECOND_STAGE_RECIPE_GROUPS),
        )

    TEMPLATE_TYPE = "CRS_RECIPE_IMPORT_EXPORT_V1"
    PENDING_DIR = os.path.join("recipe_imports", "pending")
    MAX_PENDING_UPLOAD_BYTES = _configured_max_upload_bytes()

    RECIPE_INFO_SHEET = "Recipe_Info"
    PARAMETERS_SHEET = "Parameters"
    PHASE_SHEET = "Phase_Control"
    LISTS_SHEET = "Lists"
    README_SHEET = "README"

    PARAMETER_HEADERS = [
        "Tag Index",
        "PLC Array Index",
        "Parameter Name",
        "Parameter Class",
        "Unit",
        "Data Type",
        "Min Value",
        "Max Value",
        "Default Value",
        "Parameter Value",
        "Is Modified",
        "English Memo",
        "Used",
        "Parameter Definition ID",
    ]

    LEGACY_PARAMETER_HEADERS = [
        "Name",
        "Unit",
        "Class",
        "Old Value",
        "Value",
        "Max. Value",
        "Min. Value",
        "Default Value",
        "Tag Index",
        "Memo",
        "Status",
    ]

    PHASE_HEADERS = [
        "Line No",
        "Phase Control Name",
        "Phase Control ID",
        "Stop Option",
        "Position Option",
        "Sequence No",
        "Description",
        "Stage Type",
    ]

    @staticmethod
    def _safe_file_part(value):
        value = str(value or "CRS").strip()
        value = re.sub(r"[^A-Za-z0-9_.-]+", "_", value)
        return value[:80] or "CRS"

    @staticmethod
    def _now_stamp():
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    @staticmethod
    def _get_request_context(request_obj=None):
        if not request_obj:
            return {}
        forwarded_for = request_obj.headers.get("X-Forwarded-For")
        client_ip = forwarded_for.split(",")[0].strip() if forwarded_for else request_obj.remote_addr
        return {
            "client_ip": client_ip,
            "workstation_name": (
                request_obj.headers.get("X-Workstation-Name")
                or request_obj.headers.get("X-Client-Workstation")
                or request_obj.headers.get("X-Forwarded-Host")
                or request_obj.host
            ),
            "user_agent": request_obj.headers.get("User-Agent", ""),
            "forwarded_for": forwarded_for,
            "request_host": request_obj.host,
        }

    @staticmethod
    def _apply_sheet_style(ws, freeze="A2"):
        header_fill = PatternFill("solid", fgColor="243247")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.freeze_panes = freeze
        ws.sheet_view.showGridLines = False

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 8
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells[:80]:
                value = cell.value
                if value is not None:
                    max_length = max(max_length, min(len(str(value)), 45))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 36)

    @staticmethod
    def _style_info_sheet(ws):
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 45
        title_fill = PatternFill("solid", fgColor="6F2DA8")
        sub_fill = PatternFill("solid", fgColor="F4EDFB")
        thin = Side(style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws["A1"].fill = title_fill
        ws["A1"].font = Font(color="FFFFFF", bold=True)
        ws["B1"].fill = title_fill
        ws["B1"].font = Font(color="FFFFFF", bold=True)
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, 1).fill = sub_fill
            ws.cell(row_idx, 1).font = Font(bold=True)
        ws.freeze_panes = "A2"

    @staticmethod
    def _create_base_workbook():
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        return wb

    @staticmethod
    def _write_readme(wb):
        ws = wb.create_sheet(RecipeExcelImportExportManager.README_SHEET)
        rows = [
            ["CRS Recipe Excel Template", "Purpose"],
            ["Template Type", RecipeExcelImportExportManager.TEMPLATE_TYPE],
            ["Editable fields", "Recipe_Info: Recipe Code / GT Code, Recipe Name, Version, Is Test Only. Parameters: Parameter Value and safe non-PLC details (name/unit/min/max/default/memo/used) when enabled. Phase_Control is exported for reference/manual setup but is not imported."],
            ["Safety rule", "Imported recipes are saved as DRAFT. They must be reviewed/released before production use."],
            ["Do not change", "Tag Index, PLC Array Index, and Parameter Definition ID are protected PLC mapping fields. Phase_Control rows are reference only during import."],
            ["Validation", "All used parameters for selected machine/stage must be present. Values must be numeric and within min/max where limits are configured. Phase rows are defaulted during import and must be selected manually before PLC download."],
        ]
        for row in rows:
            ws.append(row)
        RecipeExcelImportExportManager._style_info_sheet(ws)
        return ws

    @staticmethod
    def _write_recipe_info(wb, recipe=None, target=None, exported_by=None, blank=False):
        ws = wb.create_sheet(RecipeExcelImportExportManager.RECIPE_INFO_SHEET)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        source = recipe or target or {}
        rows = [
            ["Field", "Value"],
            ["Template Type", RecipeExcelImportExportManager.TEMPLATE_TYPE],
            ["Exported At", now],
            ["Exported By", exported_by or ""],
            ["Recipe ID", "" if blank else source.get("id", "")],
            ["Recipe Code / GT Code", source.get("recipe_code", "")],
            ["Recipe Name", source.get("recipe_name", "")],
            ["Version", source.get("version", 1)],
            ["Status", "DRAFT"],
            ["Machine ID", source.get("machine_id", "")],
            ["Machine Code", source.get("machine_code", "")],
            ["Stage ID", source.get("stage_id", "")],
            ["Stage Type", source.get("stage_type", "")],
            ["Is Test Only", source.get("is_test_only", 0)],
        ]
        for row in rows:
            ws.append(row)
        RecipeExcelImportExportManager._style_info_sheet(ws)
        return ws

    @staticmethod
    def _write_parameters(wb, parameter_rows):
        ws = wb.create_sheet(RecipeExcelImportExportManager.PARAMETERS_SHEET)
        ws.append(RecipeExcelImportExportManager.PARAMETER_HEADERS)
        for row in parameter_rows:
            ws.append([
                row.get("tag_index"),
                row.get("plc_array_index"),
                row.get("parameter_name"),
                row.get("parameter_class"),
                row.get("unit"),
                row.get("datatype"),
                row.get("min_value"),
                row.get("max_value"),
                row.get("default_value"),
                row.get("parameter_value"),
                row.get("is_modified", 0),
                row.get("english_memo"),
                row.get("used", 1),
                row.get("parameter_definition_id") or row.get("id"),
            ])
        RecipeExcelImportExportManager._apply_sheet_style(ws, freeze="A2")
        for col in ["A", "B", "G", "H", "I", "J", "K", "M", "N"]:
            for cell in ws[col][1:]:
                if cell.value is not None:
                    cell.number_format = "0.########"
        return ws

    @staticmethod
    def _write_phase_control(wb, phase_rows):
        ws = wb.create_sheet(RecipeExcelImportExportManager.PHASE_SHEET)
        ws.append(RecipeExcelImportExportManager.PHASE_HEADERS + [
            "Phase Group Code",
            "Phase Group Name",
        ])
        for row in phase_rows:
            ws.append([
                row.get("line_no"),
                row.get("phase_control_name"),
                row.get("phase_control_id"),
                row.get("stop_option"),
                row.get("position_option"),
                row.get("sequence_no"),
                row.get("description"),
                row.get("stage_type"),
                row.get("phase_group_code") or "MAIN",
                row.get("phase_group_name") or "Phase Control",
            ])
        RecipeExcelImportExportManager._apply_sheet_style(ws, freeze="A2")
        for col in ["A", "C", "F"]:
            for cell in ws[col][1:]:
                if cell.value is not None:
                    cell.number_format = "0"
        yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        ws.add_data_validation(yn)
        yn.add(f"D2:E{max(ws.max_row, 2)}")
        return ws

    @staticmethod
    def _write_lists(wb, stage_type):
        ws = wb.create_sheet(RecipeExcelImportExportManager.LISTS_SHEET)
        ws.append(["Phase Control ID", "Phase Control Name", "Stage Type", "Description"])
        conn = get_connection()
        cur = conn.cursor()
        phase_group_filter_sql, phase_group_filter_params = (
            RecipeExcelImportExportManager._phase_group_filter_sql(
                stage_type,
                "phase_group_code",
            )
        )
        cur.execute(
            f"""
            SELECT id, phase_control_name, stage_type, description
            FROM phase_control_master
            WHERE active = 1
                AND stage_type = ?
                {phase_group_filter_sql}
            ORDER BY
                CASE COALESCE(phase_group_code, 'MAIN')
                    WHEN 'MAIN' THEN 0
                    WHEN 'CAP_STRIP_SIDE' THEN 1
                    WHEN 'BT_SIDE' THEN 2
                    ELSE 99
                END,
                display_order,
                phase_control_name
            """,
            tuple([stage_type] + phase_group_filter_params),
        )
        for row in cur.fetchall():
            ws.append([row["id"], row["phase_control_name"], row["stage_type"], row["description"]])
        conn.close()
        RecipeExcelImportExportManager._apply_sheet_style(ws, freeze="A2")
        ws.sheet_state = "hidden"
        return ws

    @staticmethod
    def get_template_targets():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                m.id AS machine_id,
                m.machine_code,
                s.id AS stage_id,
                s.stage_type,
                COALESCE(m.description, '') AS machine_description,
                COALESCE(s.description, '') AS stage_description
            FROM tbm_machines m
            INNER JOIN machine_stages s ON s.machine_id = m.id
            WHERE COALESCE(m.active, 1) = 1 AND COALESCE(s.active, 1) = 1
            ORDER BY m.machine_code, s.stage_type
            """
        )
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows

    @staticmethod
    def _get_recipe_export_data(recipe_id):
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT r.*, m.machine_code, s.stage_type
            FROM recipes r
            LEFT JOIN tbm_machines m ON m.id = r.machine_id
            LEFT JOIN machine_stages s ON s.id = r.stage_id
            WHERE r.id = ?
            """,
            (recipe_id,),
        )
        recipe = cur.fetchone()
        if not recipe:
            conn.close()
            raise ValueError("Recipe not found.")
        recipe = dict(recipe)

        cur.execute(
            """
            SELECT
                pd.id AS parameter_definition_id,
                pd.tag_index,
                pd.plc_array_index,
                pd.parameter_name,
                pd.parameter_class,
                pd.unit,
                pd.datatype,
                pd.min_value,
                pd.max_value,
                pd.default_value,
                rpv.parameter_value,
                rpv.is_modified,
                pd.english_memo,
                pd.used
            FROM recipe_parameter_values rpv
            INNER JOIN parameter_definitions pd ON pd.id = rpv.parameter_definition_id
            WHERE rpv.recipe_id = ?
            ORDER BY pd.tag_index
            """,
            (recipe_id,),
        )
        parameters = [dict(r) for r in cur.fetchall()]

        phase_group_filter_sql, phase_group_filter_params = (
            RecipeExcelImportExportManager._phase_group_filter_sql(
                recipe.get("stage_type"),
                "rpc.phase_group_code",
            )
        )
        cur.execute(
            f"""
            SELECT
                rpc.line_no,
                rpc.phase_control_id,
                pcm.phase_control_name,
                rpc.stop_option,
                rpc.position_option,
                rpc.sequence_no,
                pcm.description,
                pcm.stage_type,
                COALESCE(rpc.phase_group_code, 'MAIN') AS phase_group_code,
                COALESCE(rpc.phase_group_name, 'Phase Control') AS phase_group_name
            FROM recipe_phase_control rpc
            LEFT JOIN phase_control_master pcm ON pcm.id = rpc.phase_control_id
            WHERE rpc.recipe_id = ?
                {phase_group_filter_sql}
            ORDER BY
                CASE COALESCE(rpc.phase_group_code, 'MAIN')
                    WHEN 'MAIN' THEN 0
                    WHEN 'CAP_STRIP_SIDE' THEN 1
                    WHEN 'BT_SIDE' THEN 2
                    ELSE 99
                END,
                rpc.line_no
            """,
            tuple([recipe_id] + phase_group_filter_params),
        )
        phase_rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return recipe, parameters, phase_rows

    @staticmethod
    def _get_target_template_data(machine_id, stage_id):
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                m.id AS machine_id,
                m.machine_code,
                s.id AS stage_id,
                s.stage_type
            FROM tbm_machines m
            INNER JOIN machine_stages s ON s.machine_id = m.id
            WHERE m.id = ? AND s.id = ?
            """,
            (machine_id, stage_id),
        )
        target = cur.fetchone()
        if not target:
            conn.close()
            raise ValueError("Machine/stage target not found.")
        target = dict(target)

        cur.execute(
            """
            SELECT
                pd.id AS parameter_definition_id,
                pd.tag_index,
                pd.plc_array_index,
                pd.parameter_name,
                pd.parameter_class,
                pd.unit,
                pd.datatype,
                pd.min_value,
                pd.max_value,
                pd.default_value,
                pd.default_value AS parameter_value,
                0 AS is_modified,
                pd.english_memo,
                pd.used
            FROM parameter_definitions pd
            WHERE pd.machine_id = ? AND pd.stage_id = ? AND COALESCE(pd.used, 1) = 1
            ORDER BY pd.tag_index
            """,
            (machine_id, stage_id),
        )
        parameters = [dict(r) for r in cur.fetchall()]

        phase_rows = RecipeExcelImportExportManager._default_phase_rows_for_target(
            cur,
            target["stage_type"]
        )
        for phase_row in phase_rows:
            phase_row["stage_type"] = target["stage_type"]
            phase_row.setdefault("description", "")
        conn.close()
        return target, parameters, phase_rows

    @staticmethod
    def build_export_workbook(recipe_id, exported_by=None):
        recipe, parameters, phase_rows = RecipeExcelImportExportManager._get_recipe_export_data(recipe_id)
        wb = RecipeExcelImportExportManager._create_base_workbook()
        RecipeExcelImportExportManager._write_readme(wb)
        RecipeExcelImportExportManager._write_recipe_info(wb, recipe=recipe, exported_by=exported_by)
        RecipeExcelImportExportManager._write_parameters(wb, parameters)
        RecipeExcelImportExportManager._write_phase_control(wb, phase_rows)
        RecipeExcelImportExportManager._write_lists(wb, recipe.get("stage_type"))
        return wb, recipe

    @staticmethod
    def build_blank_template_workbook(machine_id, stage_id, exported_by=None):
        target, parameters, phase_rows = RecipeExcelImportExportManager._get_target_template_data(machine_id, stage_id)
        target.update({
            "recipe_code": "",
            "recipe_name": "",
            "version": 1,
            "status": "DRAFT",
            "is_test_only": 0,
        })
        wb = RecipeExcelImportExportManager._create_base_workbook()
        RecipeExcelImportExportManager._write_readme(wb)
        RecipeExcelImportExportManager._write_recipe_info(wb, target=target, exported_by=exported_by, blank=True)
        RecipeExcelImportExportManager._write_parameters(wb, parameters)
        RecipeExcelImportExportManager._write_phase_control(wb, phase_rows)
        RecipeExcelImportExportManager._write_lists(wb, target.get("stage_type"))
        return wb, target

    @staticmethod
    def workbook_to_bytes(wb):
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def export_filename(recipe):
        return (
            f"CRS_RECIPE_{RecipeExcelImportExportManager._safe_file_part(recipe.get('recipe_code'))}"
            f"_V{recipe.get('version')}_{RecipeExcelImportExportManager._now_stamp()}.xlsx"
        )

    @staticmethod
    def template_filename(target):
        return (
            f"CRS_RECIPE_IMPORT_TEMPLATE_{RecipeExcelImportExportManager._safe_file_part(target.get('machine_code'))}"
            f"_{RecipeExcelImportExportManager._safe_file_part(target.get('stage_type'))}"
            f"_{RecipeExcelImportExportManager._now_stamp()}.xlsx"
        )

    @staticmethod
    def save_pending_upload(upload_file):
        if not upload_file or not upload_file.filename:
            raise ValueError("Select an Excel file to import.")
        if not upload_file.filename.lower().endswith(".xlsx"):
            raise ValueError("Only .xlsx files are supported.")
        os.makedirs(RecipeExcelImportExportManager.PENDING_DIR, exist_ok=True)
        token = uuid.uuid4().hex
        path = os.path.join(RecipeExcelImportExportManager.PENDING_DIR, f"{token}.xlsx")
        upload_file.save(path)
        if os.path.getsize(path) > RecipeExcelImportExportManager.MAX_PENDING_UPLOAD_BYTES:
            try:
                os.remove(path)
            except OSError:
                pass
            raise ValueError("Uploaded Excel file is too large.")
        return token, path

    @staticmethod
    def pending_path(token):
        token = re.sub(r"[^a-fA-F0-9]", "", str(token or ""))
        if not token:
            raise ValueError("Invalid import token.")
        path = os.path.join(RecipeExcelImportExportManager.PENDING_DIR, f"{token}.xlsx")
        if not os.path.exists(path):
            raise ValueError("Pending import file not found. Upload the template again.")
        return path

    @staticmethod
    def _read_info_sheet(wb):
        if RecipeExcelImportExportManager.RECIPE_INFO_SHEET not in wb.sheetnames:
            raise ValueError("Recipe_Info sheet missing.")
        ws = wb[RecipeExcelImportExportManager.RECIPE_INFO_SHEET]
        result = {}
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            key, value = row
            if key is None:
                continue
            result[str(key).strip().lower()] = value
        return result

    @staticmethod
    def _info_value(info, *names, default=None):
        for name in names:
            key = str(name).strip().lower()
            if key in info and info[key] not in (None, ""):
                return info[key]
        return default

    @staticmethod
    def _sheet_rows(wb, sheet_name, expected_headers):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{sheet_name} sheet missing.")
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        missing = [h for h in expected_headers if h not in headers]
        if missing:
            raise ValueError(f"{sheet_name} missing column(s): {', '.join(missing)}")
        header_index = {h: headers.index(h) for h in headers if h}
        rows = []
        for excel_row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(value is None for value in row):
                continue
            item = {h: row[header_index[h]] if header_index[h] < len(row) else None for h in expected_headers}
            for header in headers:
                if header and header not in item:
                    item[header] = row[header_index[header]] if header_index[header] < len(row) else None
            item["__excel_row__"] = excel_row_no
            rows.append(item)
        return rows

    @staticmethod
    def _normalize_excel_heading(value):
        return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())

    @staticmethod
    def _is_numeric_like(value):
        if value in (None, ""):
            return False
        try:
            float(value)
            return True
        except Exception:
            return False

    @staticmethod
    def _legacy_used_flag(value):
        if value in (None, ""):
            return 1
        text = str(value).strip().upper()
        if text in {"N", "NO", "0", "FALSE", "NOT USED", "UNUSED"}:
            return 0
        return 1

    @staticmethod
    def _legacy_metadata(wb):
        metadata = {
            "recipe_code": "",
            "recipe_name": "",
        }
        for ws in wb.worksheets:
            max_row = min(ws.max_row or 1, 10)
            for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
                row_values = list(row or [])
                for idx, cell_value in enumerate(row_values):
                    label = str(cell_value or "").strip().rstrip(":").lower()
                    next_value = row_values[idx + 1] if idx + 1 < len(row_values) else None
                    if label == "recipe name" and next_value not in (None, ""):
                        metadata["recipe_name"] = str(next_value).strip()
                    if label == "product code" and next_value not in (None, ""):
                        metadata["recipe_code"] = str(next_value).strip().upper()
        if not metadata["recipe_code"] and metadata["recipe_name"]:
            metadata["recipe_code"] = RecipeExcelImportExportManager._safe_file_part(
                metadata["recipe_name"]
            ).upper()
        if not metadata["recipe_name"] and metadata["recipe_code"]:
            metadata["recipe_name"] = metadata["recipe_code"]
        return metadata

    @staticmethod
    def _legacy_header_map(row_values):
        aliases = {
            "name": "Parameter Name",
            "parametername": "Parameter Name",
            "unit": "Unit",
            "class": "Parameter Class",
            "parameterclass": "Parameter Class",
            "oldvalue": "Old Value",
            "value": "Parameter Value",
            "maxvalue": "Max Value",
            "maximumvalue": "Max Value",
            "minvalue": "Min Value",
            "minimumvalue": "Min Value",
            "defaultvalue": "Default Value",
            "tagindex": "Tag Index",
            "tag": "Tag Index",
            "memo": "English Memo",
            "englishmemo": "English Memo",
            "status": "Used",
            "used": "Used",
        }
        header_map = {}
        for idx, value in enumerate(row_values):
            key = RecipeExcelImportExportManager._normalize_excel_heading(value)
            canonical = aliases.get(key)
            if canonical:
                header_map[canonical] = idx
        required = {"Parameter Name", "Parameter Value", "Tag Index"}
        return header_map if required.issubset(set(header_map)) else None

    @staticmethod
    def _read_legacy_parameter_rows(wb):
        """Read the tyre-machine recipe Excel format used by existing plant files.

        Supported layout example:
        Recipe Name / Product Code in top rows, then columns:
        Name, Unit, Class, Old Value, Value, Max. Value, Min. Value,
        Default Value, Tag Index, Memo, Status.
        """
        for ws in wb.worksheets:
            header_map = None
            header_row_no = None
            max_scan_row = min(ws.max_row or 1, 25)
            for row_no in range(1, max_scan_row + 1):
                values = [cell.value for cell in ws[row_no]]
                possible_map = RecipeExcelImportExportManager._legacy_header_map(values)
                if possible_map:
                    header_map = possible_map
                    header_row_no = row_no
                    break

            if header_map:
                rows = []
                for excel_row_no, row in enumerate(ws.iter_rows(min_row=header_row_no + 1, values_only=True), start=header_row_no + 1):
                    if all(value is None for value in row):
                        continue
                    tag_idx_pos = header_map.get("Tag Index")
                    value_pos = header_map.get("Parameter Value")
                    if tag_idx_pos is None or tag_idx_pos >= len(row):
                        continue
                    if not RecipeExcelImportExportManager._is_numeric_like(row[tag_idx_pos]):
                        continue
                    if value_pos is None or value_pos >= len(row):
                        continue
                    item = {
                        "Tag Index": row[header_map["Tag Index"]],
                        "PLC Array Index": row[header_map["Tag Index"]],
                        "Parameter Name": row[header_map["Parameter Name"]] if header_map.get("Parameter Name") is not None else "",
                        "Parameter Class": row[header_map["Parameter Class"]] if header_map.get("Parameter Class") is not None and header_map["Parameter Class"] < len(row) else "",
                        "Unit": row[header_map["Unit"]] if header_map.get("Unit") is not None and header_map["Unit"] < len(row) else "",
                        "Data Type": "REAL",
                        "Min Value": row[header_map["Min Value"]] if header_map.get("Min Value") is not None and header_map["Min Value"] < len(row) else None,
                        "Max Value": row[header_map["Max Value"]] if header_map.get("Max Value") is not None and header_map["Max Value"] < len(row) else None,
                        "Default Value": row[header_map["Default Value"]] if header_map.get("Default Value") is not None and header_map["Default Value"] < len(row) else None,
                        "Parameter Value": row[header_map["Parameter Value"]],
                        "Is Modified": 1,
                        "English Memo": row[header_map["English Memo"]] if header_map.get("English Memo") is not None and header_map["English Memo"] < len(row) else "",
                        "Used": RecipeExcelImportExportManager._legacy_used_flag(row[header_map["Used"]] if header_map.get("Used") is not None and header_map["Used"] < len(row) else 1),
                        "Parameter Definition ID": None,
                        "__excel_row__": excel_row_no,
                    }
                    rows.append(item)
                if rows:
                    return rows

        # Fallback for headerless copied data sheets with fixed plant columns A:K.
        for ws in wb.worksheets:
            rows = []
            for excel_row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not row or len(row) < 9:
                    continue
                if not RecipeExcelImportExportManager._is_numeric_like(row[8]):
                    continue
                if not RecipeExcelImportExportManager._is_numeric_like(row[4]):
                    continue
                item = {
                    "Tag Index": row[8],
                    "PLC Array Index": row[8],
                    "Parameter Name": row[0],
                    "Parameter Class": row[2] if len(row) > 2 else "",
                    "Unit": row[1] if len(row) > 1 else "",
                    "Data Type": "REAL",
                    "Min Value": row[6] if len(row) > 6 else None,
                    "Max Value": row[5] if len(row) > 5 else None,
                    "Default Value": row[7] if len(row) > 7 else None,
                    "Parameter Value": row[4],
                    "Is Modified": 1,
                    "English Memo": row[9] if len(row) > 9 else "",
                    "Used": RecipeExcelImportExportManager._legacy_used_flag(row[10] if len(row) > 10 else 1),
                    "Parameter Definition ID": None,
                    "__excel_row__": excel_row_no,
                }
                rows.append(item)
            if rows:
                return rows
        return []

    @staticmethod
    def _default_parameter_value(definition):
        value = definition.get("default_value")
        if value in (None, ""):
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    @staticmethod
    def _optional_excel_float(value, fallback=None):
        if value in (None, ""):
            return fallback
        try:
            return float(value)
        except Exception:
            return fallback

    @staticmethod
    def _optional_excel_int(value, fallback=1):
        if value in (None, ""):
            return fallback
        try:
            return int(float(value))
        except Exception:
            return fallback

    @staticmethod
    def _update_parameter_master_non_plc_details(cur, item):
        """Update safe non-PLC parameter definition fields from Excel.

        PLC mapping fields are intentionally excluded:
        - tag_index
        - plc_array_index
        - parameter definition id
        """
        detail = item.get("excel_detail") or {}
        if not detail:
            return 0

        definition = item["definition"]
        parameter_id = int(definition["id"])

        new_values = {
            "parameter_name": str(detail.get("Parameter Name") or definition.get("parameter_name") or "").strip(),
            "parameter_class": str(detail.get("Parameter Class") or definition.get("parameter_class") or "").strip(),
            "unit": str(detail.get("Unit") or definition.get("unit") or "").strip(),
            "datatype": str(detail.get("Data Type") or definition.get("datatype") or "REAL").strip().upper(),
            "min_value": RecipeExcelImportExportManager._optional_excel_float(detail.get("Min Value"), definition.get("min_value")),
            "max_value": RecipeExcelImportExportManager._optional_excel_float(detail.get("Max Value"), definition.get("max_value")),
            "default_value": RecipeExcelImportExportManager._optional_excel_float(detail.get("Default Value"), definition.get("default_value")),
            "english_memo": str(detail.get("English Memo") or definition.get("english_memo") or "").strip(),
            "used": RecipeExcelImportExportManager._optional_excel_int(detail.get("Used"), definition.get("used") or 1),
        }

        if not new_values["parameter_name"]:
            new_values["parameter_name"] = str(definition.get("parameter_name") or f"Tag {definition.get('tag_index')}").strip()

        changed = 0
        for field, new_value in new_values.items():
            old_value = definition.get(field)
            if str(old_value if old_value is not None else "") != str(new_value if new_value is not None else ""):
                changed += 1

        if not changed:
            return 0

        cur.execute(
            """
            UPDATE parameter_definitions
            SET
                parameter_name = ?,
                parameter_class = ?,
                unit = ?,
                datatype = ?,
                min_value = ?,
                max_value = ?,
                default_value = ?,
                english_memo = ?,
                used = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                new_values["parameter_name"],
                new_values["parameter_class"],
                new_values["unit"],
                new_values["datatype"],
                new_values["min_value"],
                new_values["max_value"],
                new_values["default_value"],
                new_values["english_memo"],
                new_values["used"],
                parameter_id,
            ),
        )
        return 1

    @staticmethod
    def _append_missing_parameter_defaults(
        definitions,
        seen_defs,
        parsed_parameters,
        warnings,
        mark_missing_parameters_not_used=False,
    ):
        missing_defs = [r for r in definitions if r["id"] not in seen_defs]
        for definition in missing_defs:
            excel_detail = None
            if mark_missing_parameters_not_used:
                excel_detail = {
                    "Parameter Name": definition.get("parameter_name"),
                    "Parameter Class": definition.get("parameter_class"),
                    "Unit": definition.get("unit"),
                    "Data Type": definition.get("datatype"),
                    "Min Value": definition.get("min_value"),
                    "Max Value": definition.get("max_value"),
                    "Default Value": definition.get("default_value"),
                    "English Memo": definition.get("english_memo"),
                    "Used": 0,
                }
            parsed_parameters.append({
                "definition": definition,
                "value": RecipeExcelImportExportManager._default_parameter_value(definition),
                "excel_detail": excel_detail,
            })
        if missing_defs:
            if mark_missing_parameters_not_used:
                warnings.append(
                    f"{len(missing_defs)} active parameter row(s) were not present in Excel. CRS will mark them Not Used and keep default values for traceability."
                )
            else:
                warnings.append(
                    f"{len(missing_defs)} parameter row(s) were not present in Excel. CRS inserted master default values for those missing rows so PLC index mapping remains complete."
                )
        return missing_defs

    @staticmethod
    def _default_phase_rows_for_target(cur, stage_type):
        if str(stage_type or "").upper() == "SECOND_STAGE":
            group_specs = [
                ("CAP_STRIP_SIDE", "Cap Strip Side"),
                ("BT_SIDE", "B&T Side"),
            ]
            phase_rows = []
            for group_code, group_name in group_specs:
                cur.execute(
                    """
                    SELECT id, phase_control_name
                    FROM phase_control_master
                    WHERE active = 1 AND stage_type = ? AND phase_group_code = ? AND phase_control_name = 'Empty Phase'
                    ORDER BY display_order, id
                    LIMIT 1
                    """,
                    (stage_type, group_code),
                )
                empty = cur.fetchone()
                if not empty:
                    cur.execute(
                        """
                        SELECT id, phase_control_name
                        FROM phase_control_master
                        WHERE active = 1 AND stage_type = ? AND phase_group_code = ?
                        ORDER BY display_order DESC, id DESC
                        LIMIT 1
                        """,
                        (stage_type, group_code),
                    )
                    empty = cur.fetchone()
                phase_id = int(empty["id"]) if empty else None
                phase_name = empty["phase_control_name"] if empty else "Empty Phase"
                for line_no in range(1, 7):
                    phase_rows.append({
                        "line_no": line_no,
                        "phase_control_id": phase_id,
                        "phase_control_name": phase_name,
                        "stop_option": "No",
                        "position_option": "No",
                        "sequence_no": line_no,
                        "phase_group_code": group_code,
                        "phase_group_name": group_name,
                    })
            return phase_rows

        group_code = "MAIN"
        group_name = "Phase Control"
        cur.execute(
            """
            SELECT id, phase_control_name
            FROM phase_control_master
            WHERE active = 1
                AND stage_type = ?
                AND phase_group_code = ?
                AND phase_control_name = 'Empty Phase'
            ORDER BY display_order, id
            LIMIT 1
            """,
            (
                stage_type,
                group_code,
            ),
        )
        empty = cur.fetchone()
        if not empty:
            cur.execute(
                """
                SELECT id, phase_control_name
                FROM phase_control_master
                WHERE active = 1 AND stage_type = ? AND phase_control_name = 'Empty Phase'
                ORDER BY display_order, id
                LIMIT 1
                """,
                (stage_type,)
            )
            empty = cur.fetchone()
        phase_id = int(empty["id"]) if empty else None
        phase_name = empty["phase_control_name"] if empty else "Empty Phase"
        return [
            {
                "line_no": line_no,
                "phase_control_id": phase_id,
                "phase_control_name": phase_name,
                "stop_option": "No",
                "position_option": "No",
                "sequence_no": line_no,
                "phase_group_code": group_code,
                "phase_group_name": group_name,
            }
            for line_no in range(1, 13)
        ]

    @staticmethod
    def _to_int(value, label, errors, row_no=None, required=True):
        if value in (None, ""):
            if required:
                errors.append(f"{label} is required" + (f" at row {row_no}" if row_no else ""))
            return None
        try:
            return int(float(value))
        except Exception:
            errors.append(f"{label} must be an integer" + (f" at row {row_no}" if row_no else ""))
            return None

    @staticmethod
    def _to_float(value, label, errors, row_no=None, required=True):
        if value in (None, ""):
            if required:
                errors.append(f"{label} is required" + (f" at row {row_no}" if row_no else ""))
            return None
        try:
            return float(value)
        except Exception:
            errors.append(f"{label} must be numeric" + (f" at row {row_no}" if row_no else ""))
            return None

    @staticmethod
    def _truthy(value):
        return str(value or "").strip().lower() in {"1", "true", "yes", "on", "checked"}

    @staticmethod
    def _normalize_import_mode(value):
        value = str(value or "create_new").strip().lower()
        if value in {"update", "update_existing", "existing", "update_recipe"}:
            return "update_existing"
        return "create_new"

    @staticmethod
    def get_import_recipe_options(machine_id=None, stage_id=None):
        """Return recipes that can be selected for existing-recipe Excel update.

        Historical released recipes are still returned with status text for
        visibility, but preview/confirm blocks updates to HISTORY_RELEASED.
        """
        conn = get_connection()
        cur = conn.cursor()
        conditions = []
        params = []
        if machine_id:
            conditions.append("r.machine_id = ?")
            params.append(machine_id)
        if stage_id:
            conditions.append("r.stage_id = ?")
            params.append(stage_id)
        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""
        cur.execute(
            f"""
            SELECT
                r.id,
                r.machine_id,
                r.stage_id,
                r.recipe_code,
                r.recipe_name,
                r.version,
                r.status,
                COALESCE(r.is_test_only, 0) AS is_test_only,
                m.machine_code,
                s.stage_type,
                CASE
                    WHEN r.status = 'RELEASED'
                         AND r.version = (
                             SELECT MAX(x.version)
                             FROM recipes x
                             WHERE x.machine_id = r.machine_id
                               AND x.stage_id = r.stage_id
                               AND UPPER(x.recipe_code) = UPPER(r.recipe_code)
                               AND x.status = 'RELEASED'
                         ) THEN 'CURRENT_RELEASED'
                    WHEN r.status = 'RELEASED' THEN 'HISTORY_RELEASED'
                    ELSE COALESCE(r.status, '')
                END AS version_usage_status
            FROM recipes r
            LEFT JOIN tbm_machines m ON m.id = r.machine_id
            LEFT JOIN machine_stages s ON s.id = r.stage_id
            {where_clause}
            ORDER BY m.machine_code, s.stage_type, r.recipe_code, r.version DESC, r.id DESC
            """,
            tuple(params),
        )
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows

    @staticmethod
    def _get_recipe_update_target(cur, recipe_id):
        if not recipe_id:
            return None
        cur.execute(
            """
            SELECT
                r.*,
                m.machine_code,
                s.stage_type,
                CASE
                    WHEN r.status = 'RELEASED'
                         AND r.version = (
                             SELECT MAX(x.version)
                             FROM recipes x
                             WHERE x.machine_id = r.machine_id
                               AND x.stage_id = r.stage_id
                               AND UPPER(x.recipe_code) = UPPER(r.recipe_code)
                               AND x.status = 'RELEASED'
                         ) THEN 'CURRENT_RELEASED'
                    WHEN r.status = 'RELEASED' THEN 'HISTORY_RELEASED'
                    ELSE COALESCE(r.status, '')
                END AS version_usage_status
            FROM recipes r
            LEFT JOIN tbm_machines m ON m.id = r.machine_id
            LEFT JOIN machine_stages s ON s.id = r.stage_id
            WHERE r.id = ?
            """,
            (recipe_id,),
        )
        row = cur.fetchone()
        return dict(row) if row else None

    @staticmethod
    def _build_effective_limits(definition, row, update_master_details):
        if update_master_details:
            min_v = RecipeExcelImportExportManager._optional_excel_float(
                row.get("Min Value"),
                definition.get("min_value"),
            )
            max_v = RecipeExcelImportExportManager._optional_excel_float(
                row.get("Max Value"),
                definition.get("max_value"),
            )
        else:
            min_v = definition.get("min_value")
            max_v = definition.get("max_value")
        return min_v, max_v

    @staticmethod
    def _current_value_map(cur, recipe_id):
        cur.execute(
            """
            SELECT
                rpv.id AS value_id,
                rpv.parameter_definition_id,
                rpv.parameter_value
            FROM recipe_parameter_values rpv
            WHERE rpv.recipe_id = ?
            """,
            (recipe_id,),
        )
        return {int(r["parameter_definition_id"]): dict(r) for r in cur.fetchall()}

    @staticmethod
    def _values_different(old_value, new_value):
        try:
            return float(old_value) != float(new_value)
        except Exception:
            return str(old_value if old_value is not None else "") != str(new_value if new_value is not None else "")

    @staticmethod
    def preview_import(
        file_path,
        machine_id=None,
        stage_id=None,
        recipe_code_override=None,
        recipe_name_override=None,
        import_mode="create_new",
        existing_recipe_id=None,
        update_master_details=True,
        mark_missing_parameters_not_used=False,
    ):
        errors = []
        warnings = []
        legacy_format = False
        import_mode = RecipeExcelImportExportManager._normalize_import_mode(import_mode)
        update_master_details = RecipeExcelImportExportManager._truthy(update_master_details)
        mark_missing_parameters_not_used = RecipeExcelImportExportManager._truthy(
            mark_missing_parameters_not_used
        )

        requested_machine_id = machine_id
        requested_stage_id = stage_id
        recipe_code_override = str(recipe_code_override or "").strip().upper()
        recipe_name_override = str(recipe_name_override or "").strip()

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as exc:
            return {"ok": False, "errors": [f"Unable to read Excel workbook: {exc}"], "warnings": [], "summary": {}}

        phase_rows = []
        phase_sheet_row_count = 0
        try:
            info = RecipeExcelImportExportManager._read_info_sheet(wb)
            template_type = RecipeExcelImportExportManager._info_value(info, "Template Type")
            if template_type != RecipeExcelImportExportManager.TEMPLATE_TYPE:
                errors.append("Invalid template type. Download a fresh CRS recipe template/export workbook.")

            sheet_recipe_code = str(RecipeExcelImportExportManager._info_value(info, "Recipe Code / GT Code", "GT Code", default="") or "").strip().upper()
            sheet_recipe_name = str(RecipeExcelImportExportManager._info_value(info, "Recipe Name", default="") or "").strip()
            recipe_code = recipe_code_override or sheet_recipe_code
            recipe_name = recipe_name_override or sheet_recipe_name

            version = RecipeExcelImportExportManager._to_int(
                RecipeExcelImportExportManager._info_value(info, "Version", default=1),
                "Version",
                errors
            )

            sheet_machine_id = RecipeExcelImportExportManager._to_int(
                RecipeExcelImportExportManager._info_value(info, "Machine ID"),
                "Machine ID",
                errors,
                required=False
            )
            sheet_stage_id = RecipeExcelImportExportManager._to_int(
                RecipeExcelImportExportManager._info_value(info, "Stage ID"),
                "Stage ID",
                errors,
                required=False
            )

            form_machine_id = RecipeExcelImportExportManager._to_int(
                requested_machine_id,
                "Target Machine ID",
                errors,
                required=False
            )
            form_stage_id = RecipeExcelImportExportManager._to_int(
                requested_stage_id,
                "Target Stage ID",
                errors,
                required=False
            )

            machine_id = form_machine_id or sheet_machine_id
            stage_id = form_stage_id or sheet_stage_id

            is_test_only = RecipeExcelImportExportManager._to_int(
                RecipeExcelImportExportManager._info_value(info, "Is Test Only", default=0),
                "Is Test Only",
                errors,
                required=False
            ) or 0

            parameter_rows = RecipeExcelImportExportManager._sheet_rows(
                wb,
                RecipeExcelImportExportManager.PARAMETERS_SHEET,
                RecipeExcelImportExportManager.PARAMETER_HEADERS,
            )

            if RecipeExcelImportExportManager.PHASE_SHEET in wb.sheetnames:
                try:
                    phase_rows = RecipeExcelImportExportManager._sheet_rows(
                        wb,
                        RecipeExcelImportExportManager.PHASE_SHEET,
                        RecipeExcelImportExportManager.PHASE_HEADERS,
                    )
                    phase_sheet_row_count = len(phase_rows)
                except Exception as phase_exc:
                    warnings.append(
                        "Phase_Control sheet was found but was not imported. "
                        "Recipe import is parameter-only; select phase controls manually after import. "
                        f"Technical detail: {phase_exc}"
                    )
        except Exception as exc:
            legacy_rows = RecipeExcelImportExportManager._read_legacy_parameter_rows(wb)
            if not legacy_rows:
                return {"ok": False, "errors": [str(exc)], "warnings": [], "summary": {}}
            legacy_format = True
            errors = []
            warnings.append("Legacy plant Excel format detected. CRS will import recipe values by Tag Index only; PLC index/master mapping will not be edited.")
            metadata = RecipeExcelImportExportManager._legacy_metadata(wb)
            recipe_code = str(
                recipe_code_override
                or metadata.get("recipe_code")
                or RecipeExcelImportExportManager._safe_file_part(os.path.splitext(os.path.basename(file_path))[0])
            ).strip().upper()
            recipe_name = str(
                recipe_name_override
                or metadata.get("recipe_name")
                or recipe_code
            ).strip()
            version = 1
            machine_id = RecipeExcelImportExportManager._to_int(
                requested_machine_id,
                "Target Machine ID",
                errors,
                required=False
            )
            stage_id = RecipeExcelImportExportManager._to_int(
                requested_stage_id,
                "Target Stage ID",
                errors,
                required=False
            )
            is_test_only = 0
            parameter_rows = legacy_rows
            phase_rows = []
            phase_sheet_row_count = 0

        conn = get_connection()
        cur = conn.cursor()
        target = {"machine_code": "-", "stage_type": "-"}
        parsed_parameters = []
        parsed_phase = []
        update_target = None
        changed_value_count = 0
        unchanged_value_count = 0
        missing_existing_value_count = 0
        master_detail_change_count = 0

        try:
            if import_mode == "update_existing":
                existing_recipe_id_int = RecipeExcelImportExportManager._to_int(
                    existing_recipe_id,
                    "Existing Recipe",
                    errors,
                    required=True,
                )
                if existing_recipe_id_int is not None:
                    update_target = RecipeExcelImportExportManager._get_recipe_update_target(cur, existing_recipe_id_int)
                if not update_target:
                    errors.append("Existing recipe to update was not found.")
                elif str(update_target.get("version_usage_status") or "").upper() == "HISTORY_RELEASED":
                    errors.append("Historical released recipe versions are locked and cannot be updated by Excel import.")
                else:
                    machine_id = int(update_target["machine_id"])
                    stage_id = int(update_target["stage_id"])
                    recipe_code = update_target["recipe_code"]
                    recipe_name = update_target["recipe_name"]
                    version = int(update_target["version"])
                    is_test_only = int(update_target.get("is_test_only") or 0)
                    target = {
                        "machine_code": update_target.get("machine_code") or "-",
                        "stage_type": update_target.get("stage_type") or "-",
                    }
            else:
                if not recipe_code:
                    errors.append("Recipe Code / GT Code is required. Enter it on the Import page or in Recipe_Info sheet.")
                if not recipe_name:
                    errors.append("Recipe Name is required. Enter it on the Import page or in Recipe_Info sheet.")
                if machine_id is None or stage_id is None:
                    errors.append("Target Machine / Stage is required. Select it on the Import page or keep Machine ID and Stage ID in Recipe_Info sheet.")
                if version is not None and version < 1:
                    errors.append("Version must be 1 or higher.")

            if not errors:
                cur.execute(
                    """
                    SELECT m.machine_code, s.stage_type
                    FROM tbm_machines m
                    INNER JOIN machine_stages s ON s.machine_id = m.id
                    WHERE m.id = ? AND s.id = ?
                    """,
                    (machine_id, stage_id),
                )
                target_row = cur.fetchone()
                if not target_row:
                    errors.append("Machine ID / Stage ID combination does not exist.")
                    target = {"machine_code": "-", "stage_type": "-"}
                else:
                    target = dict(target_row)

            if not errors and import_mode == "create_new":
                cur.execute(
                    """
                    SELECT id FROM recipes
                    WHERE machine_id = ? AND stage_id = ? AND UPPER(recipe_code) = UPPER(?) AND version = ?
                    """,
                    (machine_id, stage_id, recipe_code, version),
                )
                if cur.fetchone():
                    errors.append(f"Recipe {recipe_code} V{version} already exists for selected machine/stage. Use Update Existing Recipe Parameters mode or use a new GT code/version.")

            if not errors:
                cur.execute(
                    """
                    SELECT * FROM parameter_definitions
                    WHERE machine_id = ? AND stage_id = ?
                    ORDER BY tag_index
                    """,
                    (machine_id, stage_id),
                )
                definitions = [dict(r) for r in cur.fetchall()]
                active_definitions = [
                    row
                    for row in definitions
                    if int(row.get("used") if row.get("used") is not None else 1) == 1
                ]
                by_id = {int(r["id"]): r for r in definitions}
                by_tag = {int(r["tag_index"]): r for r in definitions}
                current_values = RecipeExcelImportExportManager._current_value_map(cur, int(update_target["id"])) if update_target else {}

                seen_defs = set()
                for row in parameter_rows:
                    row_no = row["__excel_row__"]
                    def_id = RecipeExcelImportExportManager._to_int(row.get("Parameter Definition ID"), "Parameter Definition ID", errors, row_no, required=False)
                    tag_index = RecipeExcelImportExportManager._to_int(row.get("Tag Index"), "Tag Index", errors, row_no)
                    definition_by_id = by_id.get(def_id) if def_id else None
                    definition_by_tag = by_tag.get(tag_index) if tag_index is not None else None
                    if (
                        definition_by_id
                        and definition_by_tag
                        and int(definition_by_id.get("id")) != int(definition_by_tag.get("id"))
                    ):
                        warnings.append(
                            f"Row {row_no}: stale Parameter Definition ID {def_id} does not match Tag Index {tag_index}. Tag Index mapping was used."
                        )
                        definition = definition_by_tag
                    else:
                        definition = definition_by_id or definition_by_tag
                    if not definition:
                        errors.append(f"Parameter definition not found for Tag Index {tag_index} at row {row_no}.")
                        continue
                    if definition["id"] in seen_defs:
                        errors.append(f"Duplicate parameter definition {definition['id']} at row {row_no}.")
                        continue
                    seen_defs.add(definition["id"])

                    excel_name = str(row.get("Parameter Name") or "").strip()
                    if excel_name and excel_name != definition["parameter_name"]:
                        if update_master_details:
                            warnings.append(
                                f"Row {row_no}: parameter name differs from current master. CRS will update safe non-PLC parameter details during confirm import."
                            )
                        else:
                            warnings.append(
                                f"Row {row_no}: parameter name differs from current master. Master details update is OFF, so CRS will keep the current master name."
                            )

                    effective_min_v, effective_max_v = RecipeExcelImportExportManager._build_effective_limits(
                        definition,
                        row,
                        update_master_details,
                    )

                    row_used = RecipeExcelImportExportManager._optional_excel_int(
                        row.get("Used"),
                        definition.get("used") or 1,
                    )
                    value_label = excel_name or definition["parameter_name"]
                    value = RecipeExcelImportExportManager._to_float(
                        row.get("Parameter Value"),
                        f"Parameter Value for {value_label}",
                        errors,
                        row_no,
                        required=bool(row_used),
                    )
                    if value is None and not row_used:
                        value = RecipeExcelImportExportManager._default_parameter_value(definition)
                    if value is not None and row_used:
                        if effective_min_v is not None and value < float(effective_min_v):
                            errors.append(f"Row {row_no}: {value_label} value {value} is below minimum {effective_min_v}.")
                        if effective_max_v is not None and value > float(effective_max_v):
                            errors.append(f"Row {row_no}: {value_label} value {value} is above maximum {effective_max_v}.")

                    existing_value = current_values.get(int(definition["id"])) if update_target else None
                    if update_target:
                        if existing_value:
                            if RecipeExcelImportExportManager._values_different(existing_value.get("parameter_value"), value):
                                changed_value_count += 1
                            else:
                                unchanged_value_count += 1
                        else:
                            missing_existing_value_count += 1
                            changed_value_count += 1

                    # Estimate safe master-detail changes for preview.
                    if update_master_details:
                        detail_item = {"definition": definition, "excel_detail": row}
                        detail = detail_item["excel_detail"]
                        new_values = {
                            "parameter_name": str(detail.get("Parameter Name") or definition.get("parameter_name") or "").strip(),
                            "parameter_class": str(detail.get("Parameter Class") or definition.get("parameter_class") or "").strip(),
                            "unit": str(detail.get("Unit") or definition.get("unit") or "").strip(),
                            "datatype": str(detail.get("Data Type") or definition.get("datatype") or "REAL").strip().upper(),
                            "min_value": RecipeExcelImportExportManager._optional_excel_float(detail.get("Min Value"), definition.get("min_value")),
                            "max_value": RecipeExcelImportExportManager._optional_excel_float(detail.get("Max Value"), definition.get("max_value")),
                            "default_value": RecipeExcelImportExportManager._optional_excel_float(detail.get("Default Value"), definition.get("default_value")),
                            "english_memo": str(detail.get("English Memo") or definition.get("english_memo") or "").strip(),
                            "used": RecipeExcelImportExportManager._optional_excel_int(detail.get("Used"), definition.get("used") or 1),
                        }
                        if not new_values["parameter_name"]:
                            new_values["parameter_name"] = str(definition.get("parameter_name") or f"Tag {definition.get('tag_index')}").strip()
                        for field, new_value in new_values.items():
                            old_value = definition.get(field)
                            if str(old_value if old_value is not None else "") != str(new_value if new_value is not None else ""):
                                master_detail_change_count += 1
                                break

                    parsed_parameters.append({
                        "definition": definition,
                        "value": value,
                        "excel_detail": row,
                    })

                if import_mode == "create_new":
                    missing_defs = RecipeExcelImportExportManager._append_missing_parameter_defaults(
                        active_definitions,
                        seen_defs,
                        parsed_parameters,
                        warnings,
                        mark_missing_parameters_not_used,
                    )
                    if update_master_details and mark_missing_parameters_not_used:
                        master_detail_change_count += len(missing_defs)
                else:
                    missing_count = max(0, len(active_definitions) - len(seen_defs))
                    if missing_count:
                        warnings.append(
                            f"Existing-recipe update mode: {missing_count} parameter row(s) were not present in Excel and will be left unchanged."
                        )
                    if missing_existing_value_count:
                        warnings.append(
                            f"{missing_existing_value_count} Excel row(s) do not currently have recipe value records; CRS will insert those value records during confirm update."
                        )

                parsed_phase = RecipeExcelImportExportManager._default_phase_rows_for_target(
                    cur,
                    target.get("stage_type")
                ) if import_mode == "create_new" else []

                if import_mode == "create_new":
                    if phase_sheet_row_count:
                        warnings.append(
                            f"Phase_Control sheet has {phase_sheet_row_count} row(s), but recipe import is parameter-only. "
                            "CRS ignored Excel phase text/IDs and inserted default Empty Phase rows. "
                            "Select the correct phase controls manually before PLC download."
                        )
                    else:
                        warnings.append(
                            "No phase-control rows were imported. CRS inserted default Empty Phase rows. "
                            "Select the correct phase controls manually before PLC download."
                        )
                else:
                    warnings.append(
                        "Existing-recipe update mode is parameter-only. Phase_Control sheet is ignored and existing phase selections are not changed."
                    )
        finally:
            conn.close()

        summary = {
            "recipe_code": recipe_code if 'recipe_code' in locals() else "",
            "recipe_name": recipe_name if 'recipe_name' in locals() else "",
            "version": version if 'version' in locals() else "",
            "machine_id": machine_id if 'machine_id' in locals() else "",
            "machine_code": target.get("machine_code", "-") if 'target' in locals() else "-",
            "stage_id": stage_id if 'stage_id' in locals() else "",
            "stage_type": target.get("stage_type", "-") if 'target' in locals() else "-",
            "is_test_only": is_test_only if 'is_test_only' in locals() else 0,
            "parameter_count": len(parsed_parameters) if 'parsed_parameters' in locals() else 0,
            "phase_count": len(parsed_phase) if 'parsed_phase' in locals() else 0,
            "phase_import_mode": "IGNORED_KEEP_EXISTING_PHASES" if import_mode == "update_existing" else "DEFAULT_EMPTY_PHASE_ROWS_MANUAL_SELECTION",
            "import_format": "LEGACY_PLANT_EXCEL" if 'legacy_format' in locals() and legacy_format else "CRS_TEMPLATE",
            "import_mode": import_mode,
            "mode_label": "Update Existing Recipe Parameters" if import_mode == "update_existing" else "Create New Draft Recipe",
            "existing_recipe_id": int(update_target["id"]) if update_target else "",
            "existing_recipe_status": update_target.get("version_usage_status") if update_target else "",
            "changed_value_count": changed_value_count,
            "unchanged_value_count": unchanged_value_count,
            "master_detail_change_count": master_detail_change_count,
            "update_master_details": 1 if update_master_details else 0,
            "mark_missing_parameters_not_used": 1 if mark_missing_parameters_not_used else 0,
        }
        return {
            "ok": not errors,
            "errors": errors,
            "warnings": warnings,
            "summary": summary,
            "parameters": parsed_parameters[:15] if 'parsed_parameters' in locals() else [],
            "phase_rows": parsed_phase[:24] if 'parsed_phase' in locals() else [],
            "_parsed_parameters": parsed_parameters if 'parsed_parameters' in locals() else [],
            "_parsed_phase": parsed_phase if 'parsed_phase' in locals() else [],
        }

    @staticmethod
    def import_pending_file(
        token,
        imported_by,
        user_role,
        reason=None,
        request_obj=None,
        machine_id=None,
        stage_id=None,
        recipe_code_override=None,
        recipe_name_override=None,
        import_mode="create_new",
        existing_recipe_id=None,
        update_master_details=True,
        mark_missing_parameters_not_used=False,
    ):
        file_path = RecipeExcelImportExportManager.pending_path(token)
        import_mode = RecipeExcelImportExportManager._normalize_import_mode(import_mode)
        update_master_details = RecipeExcelImportExportManager._truthy(update_master_details)
        mark_missing_parameters_not_used = RecipeExcelImportExportManager._truthy(
            mark_missing_parameters_not_used
        )
        preview = RecipeExcelImportExportManager.preview_import(
            file_path,
            machine_id=machine_id,
            stage_id=stage_id,
            recipe_code_override=recipe_code_override,
            recipe_name_override=recipe_name_override,
            import_mode=import_mode,
            existing_recipe_id=existing_recipe_id,
            update_master_details=update_master_details,
            mark_missing_parameters_not_used=mark_missing_parameters_not_used,
        )
        if not preview["ok"]:
            return False, None, preview

        summary = preview["summary"]
        parameters = preview["_parsed_parameters"]
        phase_rows = preview["_parsed_phase"]
        recipe_code = summary["recipe_code"]
        version = int(summary["version"])
        machine_id = int(summary["machine_id"])
        stage_id = int(summary["stage_id"])
        is_test_only = int(summary.get("is_test_only") or 0)
        ctx = RecipeExcelImportExportManager._get_request_context(request_obj)

        if import_mode == "update_existing":
            recipe_id = int(summary.get("existing_recipe_id") or existing_recipe_id)
            conn = get_connection()
            cur = conn.cursor()
            try:
                cur.execute("BEGIN")
                recipe = RecipeExcelImportExportManager._get_recipe_update_target(cur, recipe_id)
                if not recipe:
                    raise ValueError("Existing recipe was not found.")
                if str(recipe.get("version_usage_status") or "").upper() == "HISTORY_RELEASED":
                    raise ValueError("Historical released recipe versions are locked and cannot be updated by Excel import.")

                master_detail_update_count = 0
                if update_master_details:
                    for item in parameters:
                        master_detail_update_count += RecipeExcelImportExportManager._update_parameter_master_non_plc_details(cur, item)

                current_values = RecipeExcelImportExportManager._current_value_map(cur, recipe_id)
                changed_count = 0
                inserted_count = 0
                unchanged_count = 0
                audit_rows = []

                for item in parameters:
                    definition = item["definition"]
                    definition_id = int(definition["id"])
                    new_value = item["value"]
                    existing_value = current_values.get(definition_id)
                    if existing_value:
                        value_id = int(existing_value["value_id"])
                        old_value = existing_value.get("parameter_value")
                        if RecipeExcelImportExportManager._values_different(old_value, new_value):
                            cur.execute(
                                """
                                UPDATE recipe_parameter_values
                                SET parameter_value = ?, is_modified = 1, updated_at = CURRENT_TIMESTAMP
                                WHERE id = ?
                                """,
                                (new_value, value_id),
                            )
                            changed_count += 1
                            audit_rows.append((value_id, definition, old_value, new_value))
                        else:
                            unchanged_count += 1
                    else:
                        cur.execute(
                            """
                            INSERT INTO recipe_parameter_values
                            (recipe_id, parameter_definition_id, parameter_value, is_modified)
                            VALUES (?, ?, ?, 1)
                            """,
                            (recipe_id, definition_id, new_value),
                        )
                        value_id = cur.lastrowid
                        inserted_count += 1
                        audit_rows.append((value_id, definition, None, new_value))

                conn.commit()
            except Exception as exc:
                conn.rollback()
                conn.close()
                preview["errors"] = [f"Database update failed: {exc}"]
                preview["ok"] = False
                return False, None, preview
            conn.close()

            # Parameter value audit is written after the main transaction so the
            # update remains simple and compatible with existing audit schema.
            from database.recipe_parameter_audit_manager import RecipeParameterAuditManager
            change_source = "CURRENT_RELEASED_EDIT" if str(summary.get("existing_recipe_status") or "").upper() == "CURRENT_RELEASED" else "EXCEL_EXISTING_RECIPE_UPDATE"
            for value_id, definition, old_value, new_value in audit_rows:
                try:
                    RecipeParameterAuditManager.log_change(
                        recipe_id=recipe_id,
                        recipe_parameter_value_id=value_id,
                        parameter_definition_id=definition["id"],
                        old_value=old_value,
                        new_value=new_value,
                        changed_by=imported_by,
                        recipe_code=recipe_code,
                        recipe_version=version,
                        parameter_name=definition.get("parameter_name"),
                        tag_index=definition.get("tag_index"),
                        change_source=change_source,
                        change_reason=reason or "Existing recipe parameter update from Excel import",
                        user_role=user_role,
                        client_ip=ctx.get("client_ip"),
                        workstation_name=ctx.get("workstation_name"),
                    )
                except Exception:
                    pass

            AuditManager.log_event(
                username=imported_by,
                role=user_role,
                action="RECIPE_UPDATED_EXCEL",
                change_source="WEB_RECIPE_IMPORT_EXPORT",
                recipe_code=recipe_code,
                recipe_version=version,
                record_id=recipe_id,
                old_value=os.path.basename(file_path),
                new_value=(
                    f"mode=update_existing; changed_values={changed_count}; inserted_values={inserted_count}; "
                    f"unchanged_values={unchanged_count}; master_detail_updates={master_detail_update_count}; phase_control=ignored"
                ),
                reason=reason or "Existing recipe parameter update from Excel import",
                **ctx,
            )

            try:
                os.remove(file_path)
            except OSError:
                pass

            preview["summary"]["changed_value_count"] = changed_count
            preview["summary"]["inserted_value_count"] = inserted_count
            preview["summary"]["unchanged_value_count"] = unchanged_count
            preview["summary"]["master_detail_update_count"] = master_detail_update_count
            return True, recipe_id, preview

        conn = get_connection()
        cur = conn.cursor()
        try:
            cur.execute("BEGIN")
            cur.execute(
                """
                INSERT INTO recipes
                (machine_id, stage_id, recipe_code, recipe_name, version, status, created_by, is_test_only)
                VALUES (?, ?, ?, ?, ?, 'DRAFT', ?, ?)
                """,
                (machine_id, stage_id, recipe_code, summary["recipe_name"], version, imported_by, is_test_only),
            )
            recipe_id = cur.lastrowid
            master_detail_update_count = 0
            if update_master_details:
                for item in parameters:
                    master_detail_update_count += RecipeExcelImportExportManager._update_parameter_master_non_plc_details(cur, item)
            for item in parameters:
                definition = item["definition"]
                cur.execute(
                    """
                    INSERT INTO recipe_parameter_values
                    (recipe_id, parameter_definition_id, parameter_value, is_modified)
                    VALUES (?, ?, ?, 1)
                    """,
                    (recipe_id, definition["id"], item["value"]),
                )
            for row in phase_rows:
                cur.execute(
                    """
                    INSERT INTO recipe_phase_control
                    (recipe_id, line_no, phase_control_id, stop_option, position_option, sequence_no, phase_group_code, phase_group_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        recipe_id,
                        row["line_no"],
                        row["phase_control_id"],
                        row["stop_option"],
                        row["position_option"],
                        row["sequence_no"],
                        row.get("phase_group_code") or "MAIN",
                        row.get("phase_group_name") or "Phase Control",
                    ),
                )
            conn.commit()
        except Exception as exc:
            conn.rollback()
            conn.close()
            preview["errors"] = [f"Database import failed: {exc}"]
            preview["ok"] = False
            return False, None, preview
        conn.close()

        RecipeStatusHistoryManager.add_history(
            recipe_id=recipe_id,
            recipe_code=recipe_code,
            old_status="IMPORT",
            new_status="DRAFT",
            changed_by=imported_by,
            remarks=reason or "Recipe imported from Excel template",
        )

        AuditManager.log_event(
            username=imported_by,
            role=user_role,
            action="RECIPE_IMPORTED_EXCEL",
            change_source="WEB_RECIPE_IMPORT_EXPORT",
            recipe_code=recipe_code,
            recipe_version=version,
            record_id=recipe_id,
            old_value=os.path.basename(file_path),
            new_value=f"mode=create_new; parameters={len(parameters)}; phase_rows={len(phase_rows)}; master_detail_updates={locals().get('master_detail_update_count', 0)}; status=DRAFT",
            reason=reason or "Recipe imported from Excel template",
            **ctx,
        )

        try:
            os.remove(file_path)
        except OSError:
            pass

        return True, recipe_id, preview
