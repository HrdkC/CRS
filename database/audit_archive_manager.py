import os
import re
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database.database import get_connection
from database.schema_guard import require_tables


class AuditArchiveManager:
    """Audit archive and Excel export utilities for CRS traceability."""

    DEFAULT_EXPORT_LOCATIONS = [
        r"D:\\CRS_Audit_Exports",
        r"\\\\SERVER\\CRS_Audit_Exports",
    ]

    @staticmethod
    def ensure_tables():
        return require_tables({
            "audit_log_archive": {"archive_id", "original_audit_id", "archive_batch_id"},
            "audit_archive_exports": {"id", "export_type", "export_path", "row_count"},
        })

    @staticmethod
    def _audit_columns(cursor):
        cursor.execute("PRAGMA table_info(audit_log)")
        return [row[1] for row in cursor.fetchall()]

    @staticmethod
    def get_approved_export_locations():
        """
        Return plant-approved server export locations.

        Browser security does not allow a Flask web page to browse arbitrary
        server folders. Use a controlled whitelist instead. Plant can override
        using environment variable CRS_AUDIT_EXPORT_LOCATIONS with paths
        separated by semicolon, for example:
        D:\\CRS_Audit_Exports;\\SERVER\\CRS_Audit_Exports
        """
        raw = os.environ.get("CRS_AUDIT_EXPORT_LOCATIONS", "").strip()
        locations = []
        if raw:
            locations = [p.strip() for p in raw.split(";") if p.strip()]
        if not locations:
            locations = list(AuditArchiveManager.DEFAULT_EXPORT_LOCATIONS)
        seen = set()
        final = []
        for path in locations:
            key = path.lower().rstrip("\\/")
            if key not in seen:
                seen.add(key)
                final.append(path)
        return final

    @staticmethod
    def validate_export_location(export_path):
        export_path = (export_path or "").strip()
        if not export_path:
            raise ValueError("Select an approved export location.")
        approved = AuditArchiveManager.get_approved_export_locations()
        normalized = export_path.lower().rstrip("\\/")
        approved_normalized = [p.lower().rstrip("\\/") for p in approved]
        if normalized not in approved_normalized:
            raise ValueError("Export location is not in the approved plant export path list.")
        return export_path

    @staticmethod
    def _ensure_export_dir(export_path):
        export_path = AuditArchiveManager.validate_export_location(export_path)
        os.makedirs(export_path, exist_ok=True)
        return export_path

    @staticmethod
    def _safe_type(export_type):
        return ''.join(ch for ch in str(export_type or "AUDIT_EXPORT") if ch.isalnum() or ch in ('_', '-'))

    @staticmethod
    def _make_file_name(export_type):
        now_stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return f"CRS_{AuditArchiveManager._safe_type(export_type)}_{now_stamp}.xlsx"

    @staticmethod
    def build_excel_workbook(rows, sheet_title="Audit"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        headers = [
            "ID", "Timestamp", "Username", "Role", "Action", "Source",
            "Recipe Code", "Recipe Version", "Record ID", "Parameter",
            "Old Value", "New Value", "PLC", "Client IP", "Workstation",
            "User Agent", "Request Host", "Forwarded For", "Reason"
        ]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            ws.append([
                row.get("id") or row.get("original_audit_id") or row.get("archive_id"),
                row.get("timestamp"), row.get("username"), row.get("role"),
                row.get("action"), row.get("change_source"), row.get("recipe_code"),
                row.get("recipe_version"), row.get("record_id"), row.get("parameter_name"),
                row.get("old_value"), row.get("new_value"), row.get("plc_name"),
                row.get("client_ip"), row.get("workstation_name"), row.get("user_agent"),
                row.get("request_host"), row.get("forwarded_for"), row.get("reason")
            ])
        for column_cells in ws.columns:
            values = [str(c.value or "") for c in column_cells[:50]]
            width = min(55, max(12, max(len(v) for v in values) + 2))
            ws.column_dimensions[column_cells[0].column_letter].width = width
        ws.freeze_panes = "A2"
        return wb

    @staticmethod
    def build_excel_bytes(rows, export_type="AUDIT_EXPORT"):
        wb = AuditArchiveManager.build_excel_workbook(rows=rows, sheet_title=export_type)
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def get_archive_count():
        AuditArchiveManager.ensure_tables()
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM audit_log_archive")
        count = cursor.fetchone()[0]
        conn.close()
        return count

    @staticmethod
    def archive_older_than(retention_days, archived_by, export_path=None, export_excel=False, remarks=None):
        AuditArchiveManager.ensure_tables()
        try:
            retention_days = int(retention_days)
        except Exception:
            raise ValueError("Retention days must be a number.")
        if retention_days < 1:
            raise ValueError("Retention days must be at least 1.")

        if export_excel:
            export_path = AuditArchiveManager.validate_export_location(export_path)

        conn = get_connection()
        cursor = conn.cursor()
        columns = AuditArchiveManager._audit_columns(cursor)

        optional_cols = ["user_agent", "request_host", "forwarded_for"]
        select_cols = [
            "id", "username", "role", "workstation_name", "client_ip", "plc_name",
            "recipe_code", "recipe_version", "record_id", "parameter_name", "old_value",
            "new_value", "action", "change_source", "reason", "timestamp"
        ]
        for col in optional_cols:
            if col in columns and col not in select_cols:
                select_cols.insert(-1, col)

        cursor.execute(
            f"""
            SELECT {', '.join(select_cols)}
            FROM audit_log
            WHERE datetime(timestamp) < datetime('now', ?)
            ORDER BY id ASC
            """,
            (f"-{retention_days} days",)
        )
        rows = [dict(row) for row in cursor.fetchall()]
        if not rows:
            conn.close()
            return {"archived_count": 0, "export_file": None, "message": "No records eligible for archive."}

        batch_id = datetime.utcnow().strftime("AUDARCH_%Y%m%d_%H%M%S")
        for row in rows:
            cursor.execute(
                """
                INSERT INTO audit_log_archive
                (
                    original_audit_id, username, role, workstation_name, client_ip,
                    plc_name, recipe_code, recipe_version, record_id, parameter_name,
                    old_value, new_value, action, change_source, reason,
                    user_agent, request_host, forwarded_for, timestamp,
                    archived_by, archive_batch_id
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.get("id"), row.get("username"), row.get("role"),
                    row.get("workstation_name"), row.get("client_ip"), row.get("plc_name"),
                    row.get("recipe_code"), row.get("recipe_version"), row.get("record_id"),
                    row.get("parameter_name"), row.get("old_value"), row.get("new_value"),
                    row.get("action"), row.get("change_source"), row.get("reason"),
                    row.get("user_agent"), row.get("request_host"), row.get("forwarded_for"),
                    row.get("timestamp"), archived_by, batch_id,
                )
            )

        ids = [row["id"] for row in rows]
        placeholders = ",".join(["?"] * len(ids))
        cursor.execute(f"DELETE FROM audit_log WHERE id IN ({placeholders})", ids)
        conn.commit()
        conn.close()

        export_file = None
        if export_excel:
            export_file = AuditArchiveManager.export_rows_to_excel(
                rows=rows,
                export_path=export_path,
                exported_by=archived_by,
                export_type="ARCHIVE_BATCH",
                remarks=remarks or f"Archive batch {batch_id}"
            )
        return {"archived_count": len(rows), "export_file": export_file, "batch_id": batch_id}

    @staticmethod
    def export_rows_to_excel(rows, export_path, exported_by, export_type="AUDIT_EXPORT", remarks=None):
        export_path = AuditArchiveManager._ensure_export_dir(export_path)
        file_name = AuditArchiveManager._make_file_name(export_type)
        full_path = os.path.join(export_path, file_name)

        wb = AuditArchiveManager.build_excel_workbook(rows=rows, sheet_title=export_type)
        wb.save(full_path)

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO audit_archive_exports
            (export_type, export_path, file_name, row_count, exported_by, remarks)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (export_type, export_path, file_name, len(rows), exported_by, remarks)
        )
        conn.commit()
        conn.close()
        return full_path

    @staticmethod
    def get_archive_history(limit=250, username=None, role=None, action=None, change_source=None, keyword=None):
        AuditArchiveManager.ensure_tables()
        conn = get_connection()
        cursor = conn.cursor()
        conditions = []
        params = []
        if username:
            conditions.append("LOWER(username) LIKE LOWER(?)")
            params.append(f"%{username}%")
        if role:
            conditions.append("UPPER(role)=UPPER(?)")
            params.append(role)
        if action:
            conditions.append("UPPER(action)=UPPER(?)")
            params.append(action)
        if change_source:
            conditions.append("UPPER(change_source)=UPPER(?)")
            params.append(change_source)
        if keyword:
            kv = f"%{keyword}%"
            conditions.append("(LOWER(COALESCE(reason,'')) LIKE LOWER(?) OR LOWER(COALESCE(recipe_code,'')) LIKE LOWER(?) OR LOWER(COALESCE(parameter_name,'')) LIKE LOWER(?) OR LOWER(COALESCE(username,'')) LIKE LOWER(?))")
            params.extend([kv, kv, kv, kv])
        where = "WHERE " + " AND ".join(conditions) if conditions else ""
        safe_limit = max(25, min(int(limit or 250), 5000))
        cursor.execute(
            f"""
            SELECT *
            FROM audit_log_archive
            {where}
            ORDER BY timestamp DESC, archive_id DESC
            LIMIT ?
            """,
            params + [safe_limit]
        )
        rows = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return rows

    @staticmethod
    def get_recent_exports(limit=20):
        AuditArchiveManager.ensure_tables()
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT *
            FROM audit_archive_exports
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,)
        )
        rows = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return rows
