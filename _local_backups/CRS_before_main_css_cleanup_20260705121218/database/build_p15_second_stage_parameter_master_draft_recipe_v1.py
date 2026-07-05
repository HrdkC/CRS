"""
P15 Second Stage Parameter Master + Draft Recipe V1

Purpose:
- Read actual P15 SECOND_STAGE parameter data from Excel.
- Insert/update P15 SECOND_STAGE parameter_definitions.
- Ensure P15 SECOND_STAGE 3-group phase-control masters exist.
- Create first P15 SECOND_STAGE DRAFT recipe.
- Create recipe_parameter_values from parameter default values.
- Create recipe_phase_control rows from group-wise second-stage phase-control masters.

Safety:
- Creates DRAFT recipe only.
- Does not release recipe.
- Does not write PLC.
- Does not delete existing data.
- Blocks duplicate recipe_code + version + machine_id + stage_id.

Expected Excel path:
  data_imports/P15_Second_Stage_Foundation_Input_Template_V3_PhaseOnly.xlsx

Required Excel sheets:
  Recipe_Info      optional but recommended
  Parameters       required
  Phase_Control    optional; if empty, default P15 second-stage phase masters are used
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install with: pip install openpyxl"
    ) from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DB_PATH = PROJECT_ROOT / "database" / "recipe.db"
DEFAULT_EXCEL_PATH = PROJECT_ROOT / "data_imports" / "P15_Second_Stage_Foundation_Input_Template_V3_PhaseOnly.xlsx"

P15_MACHINE_CODE = "P15"
SECOND_STAGE_TYPE = "SECOND_STAGE"
DEFAULT_CREATED_BY = "admin"
DEFAULT_RECIPE_CODE = "GT_P15_SS_TEST_001"
DEFAULT_RECIPE_NAME = "P15 Second Stage Test Recipe 001"
DEFAULT_RECIPE_VERSION = 1

PHASE_GROUPS = [
    ("CAP_STRIP_SIDE", "Cap Strip Side", "P15 Second Stage Cap Strip Side phase-control group", 1),
    ("BT_SIDE", "B&T Side", "P15 Second Stage Belt and Tread side phase-control group", 2),
]

DEFAULT_PHASES = [
    ("CAP_STRIP_SIDE", "Cap Strip Side", "Apply CapStrip", "Apply CapStrip", 1),
    ("CAP_STRIP_SIDE", "Cap Strip Side", "Apply Tread", "Apply Tread", 2),
    ("BT_SIDE", "B&T Side", "Apply Belt 1", "Apply Belt 1", 1),
    ("BT_SIDE", "B&T Side", "Apply Belt 2", "Apply Belt 2", 2),
    ("BT_SIDE", "B&T Side", "Turn Table", "Turn Table", 3),
    ("BT_SIDE", "B&T Side", "Apply Tread", "Apply Tread", 4),
    ("BT_SIDE", "B&T Side", "Remove Belt Package", "Remove Belt Package", 5),
]

GROUP_NAME_BY_CODE = {code: name for code, name, _desc, _order in PHASE_GROUPS}


class ValidationError(RuntimeError):
    """Raised when input data is unsafe or incomplete."""


def normalize_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    for ch in [" ", "_", "-", ":", "/", "\\", ".", "(", ")"]:
        text = text.replace(ch, "")
    return text


def clean_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def parse_int(value: Any, field_name: str, row_no: int | None = None, default: int | None = None) -> int:
    if value is None or str(value).strip() == "":
        if default is not None:
            return default
        suffix = f" at row {row_no}" if row_no else ""
        raise ValidationError(f"Missing integer value for {field_name}{suffix}")
    try:
        return int(float(str(value).strip()))
    except Exception as exc:
        suffix = f" at row {row_no}" if row_no else ""
        raise ValidationError(f"Invalid integer for {field_name}{suffix}: {value!r}") from exc


def parse_float(value: Any, field_name: str, row_no: int | None = None, default: float | None = None) -> float | None:
    if value is None or str(value).strip() == "":
        return default
    try:
        return float(str(value).strip())
    except Exception as exc:
        suffix = f" at row {row_no}" if row_no else ""
        raise ValidationError(f"Invalid numeric value for {field_name}{suffix}: {value!r}") from exc


def parse_used(value: Any) -> int:
    if value is None or str(value).strip() == "":
        return 1
    text = str(value).strip().lower()
    if text in {"1", "yes", "y", "true", "active", "used"}:
        return 1
    if text in {"0", "no", "n", "false", "inactive", "unused"}:
        return 0
    return 1


def table_exists(cur: sqlite3.Cursor, table_name: str) -> bool:
    row = cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def get_columns(cur: sqlite3.Cursor, table_name: str) -> set[str]:
    rows = cur.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {row["name"] for row in rows}


def add_column_if_missing(cur: sqlite3.Cursor, table_name: str, column_name: str, column_sql: str) -> None:
    columns = get_columns(cur, table_name)
    if column_name not in columns:
        cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_sql}")
        print(f"Added column: {table_name}.{column_name}")


def insert_dynamic(cur: sqlite3.Cursor, table_name: str, data: dict[str, Any]) -> int:
    columns = get_columns(cur, table_name)
    filtered = {key: value for key, value in data.items() if key in columns}
    if not filtered:
        raise RuntimeError(f"No matching columns for insert into {table_name}")
    col_sql = ", ".join(filtered.keys())
    placeholder_sql = ", ".join(["?"] * len(filtered))
    cur.execute(
        f"INSERT INTO {table_name} ({col_sql}) VALUES ({placeholder_sql})",
        tuple(filtered.values()),
    )
    return int(cur.lastrowid)


def update_dynamic(cur: sqlite3.Cursor, table_name: str, row_id: int, data: dict[str, Any]) -> None:
    columns = get_columns(cur, table_name)
    filtered = {key: value for key, value in data.items() if key in columns and key != "id"}
    if not filtered:
        return
    set_sql = ", ".join([f"{key} = ?" for key in filtered.keys()])
    cur.execute(
        f"UPDATE {table_name} SET {set_sql} WHERE id = ?",
        tuple(filtered.values()) + (row_id,),
    )


def resolve_p15_second_stage(cur: sqlite3.Cursor) -> tuple[int, int]:
    machine = cur.execute(
        "SELECT id, machine_code FROM tbm_machines WHERE machine_code = ?",
        (P15_MACHINE_CODE,),
    ).fetchone()
    if not machine:
        raise ValidationError("P15 machine not found in tbm_machines.")

    stage = cur.execute(
        "SELECT id, machine_id, stage_type FROM machine_stages WHERE machine_id = ? AND stage_type = ?",
        (machine["id"], SECOND_STAGE_TYPE),
    ).fetchone()
    if not stage:
        raise ValidationError("P15 SECOND_STAGE not found in machine_stages.")

    return int(machine["id"]), int(stage["id"])


def ensure_phase_schema(cur: sqlite3.Cursor) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS phase_control_group_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            machine_stage_id INTEGER NOT NULL,
            stage_type TEXT NOT NULL,
            phase_group_code TEXT NOT NULL,
            phase_group_name TEXT NOT NULL,
            description TEXT,
            display_order INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(machine_stage_id, phase_group_code)
        )
        """
    )
    add_column_if_missing(cur, "phase_control_master", "machine_stage_id", "machine_stage_id INTEGER")
    add_column_if_missing(cur, "phase_control_master", "phase_group_code", "phase_group_code TEXT DEFAULT 'MAIN'")
    add_column_if_missing(cur, "phase_control_master", "phase_group_name", "phase_group_name TEXT DEFAULT 'Phase Control'")
    add_column_if_missing(cur, "recipe_phase_control", "phase_group_code", "phase_group_code TEXT DEFAULT 'MAIN'")
    add_column_if_missing(cur, "recipe_phase_control", "phase_group_name", "phase_group_name TEXT DEFAULT 'Phase Control'")
    add_column_if_missing(cur, "recipe_phase_control", "used", "used INTEGER DEFAULT 1")

    cur.execute("UPDATE phase_control_master SET phase_group_code='MAIN' WHERE phase_group_code IS NULL OR TRIM(phase_group_code)='' ")
    cur.execute("UPDATE phase_control_master SET phase_group_name='Phase Control' WHERE phase_group_name IS NULL OR TRIM(phase_group_name)='' ")
    cur.execute("UPDATE recipe_phase_control SET phase_group_code='MAIN' WHERE phase_group_code IS NULL OR TRIM(phase_group_code)='' ")
    cur.execute("UPDATE recipe_phase_control SET phase_group_name='Phase Control' WHERE phase_group_name IS NULL OR TRIM(phase_group_name)='' ")
    cur.execute("UPDATE recipe_phase_control SET used=1 WHERE used IS NULL")


def ensure_phase_group_masters(cur: sqlite3.Cursor, machine_stage_id: int) -> None:
    for code, name, description, display_order in PHASE_GROUPS:
        existing = cur.execute(
            "SELECT id FROM phase_control_group_master WHERE machine_stage_id=? AND phase_group_code=?",
            (machine_stage_id, code),
        ).fetchone()
        data = {
            "machine_stage_id": machine_stage_id,
            "stage_type": SECOND_STAGE_TYPE,
            "phase_group_code": code,
            "phase_group_name": name,
            "description": description,
            "display_order": display_order,
            "active": 1,
        }
        if existing:
            update_dynamic(cur, "phase_control_group_master", int(existing["id"]), data)
        else:
            insert_dynamic(cur, "phase_control_group_master", data)


def ensure_phase_master(
    cur: sqlite3.Cursor,
    machine_stage_id: int,
    phase_group_code: str,
    phase_group_name: str,
    phase_control_name: str,
    description: str,
    display_order: int,
) -> int:
    existing = cur.execute(
        """
        SELECT id
        FROM phase_control_master
        WHERE machine_stage_id=?
          AND stage_type=?
          AND phase_group_code=?
          AND phase_control_name=?
        """,
        (machine_stage_id, SECOND_STAGE_TYPE, phase_group_code, phase_control_name),
    ).fetchone()
    data = {
        "machine_stage_id": machine_stage_id,
        "stage_type": SECOND_STAGE_TYPE,
        "phase_group_code": phase_group_code,
        "phase_group_name": phase_group_name,
        "phase_control_name": phase_control_name,
        "description": description,
        "display_order": display_order,
        "active": 1,
    }
    if existing:
        update_dynamic(cur, "phase_control_master", int(existing["id"]), data)
        return int(existing["id"])
    return insert_dynamic(cur, "phase_control_master", data)


def ensure_default_phase_masters(cur: sqlite3.Cursor, machine_stage_id: int) -> None:
    for group_code, group_name, phase_name, description, display_order in DEFAULT_PHASES:
        ensure_phase_master(
            cur,
            machine_stage_id,
            group_code,
            group_name,
            phase_name,
            description,
            display_order,
        )


def read_workbook(path: Path):
    if not path.exists():
        raise FileNotFoundError(
            f"Excel input file not found: {path}\n"
            "Fill the P15 Second Stage V3 template and save it at this path, or pass --excel <path>."
        )
    return load_workbook(path, data_only=True)


def sheet_to_dict_rows(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_text(cell) for cell in rows[0]]
    normalized_headers = [normalize_key(h) for h in headers]
    result: list[dict[str, Any]] = []
    for index, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item = {normalized_headers[i]: row[i] if i < len(row) else None for i in range(len(normalized_headers)) if normalized_headers[i]}
        item["__row_no"] = index
        result.append(item)
    return result


def read_recipe_info(wb) -> dict[str, Any]:
    info = {
        "recipe_code": DEFAULT_RECIPE_CODE,
        "recipe_name": DEFAULT_RECIPE_NAME,
        "version": DEFAULT_RECIPE_VERSION,
        "created_by": DEFAULT_CREATED_BY,
    }
    if "Recipe_Info" not in wb.sheetnames:
        return info
    ws = wb["Recipe_Info"]
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        key = normalize_key(row[0])
        value = row[1]
        if not key or value is None or str(value).strip() == "":
            continue
        if key in {"recipecode", "gtcode", "gtcoderecipecode", "draftrecipecode"}:
            info["recipe_code"] = clean_text(value)
        elif key in {"recipename", "draftrecipename"}:
            info["recipe_name"] = clean_text(value)
        elif key in {"version", "recipeversion"}:
            info["version"] = parse_int(value, "version")
        elif key in {"createdby", "importedby", "user"}:
            info["created_by"] = clean_text(value, DEFAULT_CREATED_BY)
    return info


def read_parameters(wb) -> list[dict[str, Any]]:
    if "Parameters" not in wb.sheetnames:
        raise ValidationError("Excel sheet missing: Parameters")
    rows = sheet_to_dict_rows(wb["Parameters"])
    parameters: list[dict[str, Any]] = []
    seen_tag_indexes: set[int] = set()
    seen_plc_indexes: set[int] = set()
    errors: list[str] = []
    warnings: list[str] = []

    for row in rows:
        row_no = int(row["__row_no"])
        parameter_name = clean_text(row.get("parametername"))
        if not parameter_name:
            continue
        try:
            tag_index = parse_int(row.get("tagindex"), "tag_index", row_no)
            plc_array_index = parse_int(row.get("plcarrayindex"), "plc_array_index", row_no)
            min_value = parse_float(row.get("minvalue"), "min_value", row_no, default=None)
            max_value = parse_float(row.get("maxvalue"), "max_value", row_no, default=None)
            default_value = parse_float(row.get("defaultvalue"), "default_value", row_no, default=0.0)
            if default_value is None:
                default_value = 0.0
            if min_value is not None and max_value is not None and min_value > max_value:
                raise ValidationError(f"min_value > max_value at row {row_no}")
            if min_value is not None and default_value < min_value:
                raise ValidationError(f"default_value below min_value at row {row_no}: {default_value} < {min_value}")
            if max_value is not None and default_value > max_value:
                raise ValidationError(f"default_value above max_value at row {row_no}: {default_value} > {max_value}")
            if tag_index in seen_tag_indexes:
                raise ValidationError(f"Duplicate tag_index at row {row_no}: {tag_index}")
            if plc_array_index in seen_plc_indexes:
                raise ValidationError(f"Duplicate plc_array_index at row {row_no}: {plc_array_index}")
            seen_tag_indexes.add(tag_index)
            seen_plc_indexes.add(plc_array_index)

            parameters.append(
                {
                    "tag_index": tag_index,
                    "plc_array_index": plc_array_index,
                    "parameter_name": parameter_name,
                    "parameter_class": clean_text(row.get("parameterclass")),
                    "unit": clean_text(row.get("unit")),
                    "min_value": min_value,
                    "max_value": max_value,
                    "default_value": default_value,
                    "datatype": clean_text(row.get("datatype"), "REAL").upper(),
                    "english_memo": clean_text(row.get("englishmemo")),
                    "used": parse_used(row.get("used")),
                }
            )
        except ValidationError as exc:
            errors.append(str(exc))

    if not parameters:
        errors.append("No valid parameter rows found in Parameters sheet.")
    if errors:
        raise ValidationError("\n".join(errors))
    if warnings:
        print("\nWarnings:")
        for warning in warnings:
            print(f"- {warning}")
    return sorted(parameters, key=lambda x: int(x["tag_index"]))


def read_phase_rows(wb) -> list[dict[str, Any]]:
    if "Phase_Control" not in wb.sheetnames:
        return []
    rows = sheet_to_dict_rows(wb["Phase_Control"])
    phase_rows: list[dict[str, Any]] = []
    seen: set[tuple[str, int]] = set()
    auto_line_by_group: dict[str, int] = {code: 0 for code, *_ in PHASE_GROUPS}
    errors: list[str] = []

    for row in rows:
        row_no = int(row["__row_no"])
        phase_name = clean_text(row.get("phasecontrolname"))
        if not phase_name:
            continue
        used = parse_used(row.get("used"))
        if used != 1:
            continue
        group_code = clean_text(row.get("phasegroupcode")).upper()
        if not group_code:
            group_code = clean_text(row.get("groupcode")).upper()
        if group_code not in GROUP_NAME_BY_CODE:
            errors.append(f"Invalid phase_group_code at row {row_no}: {group_code!r}")
            continue
        if row.get("lineno") is None or str(row.get("lineno")).strip() == "":
            auto_line_by_group[group_code] += 1
            line_no = auto_line_by_group[group_code]
        else:
            line_no = parse_int(row.get("lineno"), "line_no", row_no)
            auto_line_by_group[group_code] = max(auto_line_by_group[group_code], line_no)
        key = (group_code, line_no)
        if key in seen:
            errors.append(f"Duplicate phase row for {group_code} line {line_no} at row {row_no}")
            continue
        seen.add(key)
        phase_rows.append(
            {
                "phase_group_code": group_code,
                "phase_group_name": GROUP_NAME_BY_CODE[group_code],
                "line_no": line_no,
                "phase_control_name": phase_name,
                "description": clean_text(row.get("description"), phase_name),
                "display_order": parse_int(row.get("displayorder"), "display_order", row_no, default=line_no),
                "used": 1,
            }
        )

    if errors:
        raise ValidationError("\n".join(errors))
    return sorted(
        phase_rows,
        key=lambda x: (list(GROUP_NAME_BY_CODE.keys()).index(x["phase_group_code"]), int(x["line_no"])),
    )


def get_default_phase_rows_from_db(cur: sqlite3.Cursor, machine_stage_id: int) -> list[dict[str, Any]]:
    rows = cur.execute(
        """
        SELECT phase_group_code, phase_group_name, phase_control_name, description, display_order
        FROM phase_control_master
        WHERE machine_stage_id=? AND stage_type=? AND active=1
        ORDER BY
            CASE phase_group_code
                WHEN 'CAP_STRIP_SIDE' THEN 1
                WHEN 'BT_SIDE' THEN 2
                WHEN 'SHAPING_SIDE' THEN 3
                ELSE 99
            END,
            display_order
        """,
        (machine_stage_id, SECOND_STAGE_TYPE),
    ).fetchall()
    result: list[dict[str, Any]] = []
    line_tracker: dict[str, int] = {}
    for row in rows:
        group_code = row["phase_group_code"]
        line_tracker[group_code] = line_tracker.get(group_code, 0) + 1
        result.append(
            {
                "phase_group_code": group_code,
                "phase_group_name": row["phase_group_name"],
                "line_no": line_tracker[group_code],
                "phase_control_name": row["phase_control_name"],
                "description": row["description"] or row["phase_control_name"],
                "display_order": line_tracker[group_code],
                "used": 1,
            }
        )
    return result


def upsert_parameter_definition(cur: sqlite3.Cursor, machine_id: int, stage_id: int, row: dict[str, Any], created_by: str) -> int:
    existing_rows = cur.execute(
        "SELECT id FROM parameter_definitions WHERE machine_id=? AND stage_id=? AND tag_index=?",
        (machine_id, stage_id, row["tag_index"]),
    ).fetchall()
    data = {
        "machine_id": machine_id,
        "stage_id": stage_id,
        "tag_index": row["tag_index"],
        "plc_array_index": row["plc_array_index"],
        "parameter_name": row["parameter_name"],
        "parameter_class": row["parameter_class"],
        "unit": row["unit"],
        "min_value": row["min_value"],
        "max_value": row["max_value"],
        "default_value": row["default_value"],
        "datatype": row["datatype"],
        "english_memo": row["english_memo"],
        "used": row["used"],
        "created_by": created_by,
        "updated_at": "CURRENT_TIMESTAMP",
    }
    # updated_at as string would be literal if inserted dynamically; use explicit SQL for update.
    data.pop("updated_at", None)

    if len(existing_rows) > 1:
        raise ValidationError(f"Multiple parameter_definitions found for tag_index {row['tag_index']} in P15 SECOND_STAGE.")
    if existing_rows:
        parameter_id = int(existing_rows[0]["id"])
        update_dynamic(cur, "parameter_definitions", parameter_id, data)
        if "updated_at" in get_columns(cur, "parameter_definitions"):
            cur.execute("UPDATE parameter_definitions SET updated_at=CURRENT_TIMESTAMP WHERE id=?", (parameter_id,))
        return parameter_id
    return insert_dynamic(cur, "parameter_definitions", data)


def create_draft_recipe(
    cur: sqlite3.Cursor,
    machine_id: int,
    stage_id: int,
    recipe_code: str,
    recipe_name: str,
    version: int,
    created_by: str,
) -> int:
    duplicate = cur.execute(
        """
        SELECT id, status
        FROM recipes
        WHERE recipe_code=? AND version=? AND machine_id=? AND stage_id=?
        """,
        (recipe_code, version, machine_id, stage_id),
    ).fetchone()
    if duplicate:
        raise ValidationError(
            f"Recipe already exists for P15 SECOND_STAGE: code={recipe_code}, version={version}, id={duplicate['id']}, status={duplicate['status']}. "
            "Use a new recipe code/version."
        )
    data = {
        "recipe_code": recipe_code,
        "recipe_name": recipe_name,
        "version": version,
        "status": "DRAFT",
        "machine_id": machine_id,
        "stage_id": stage_id,
        "created_by": created_by,
        "updated_by": created_by,
        "created_at": "CURRENT_TIMESTAMP",
        "updated_at": "CURRENT_TIMESTAMP",
    }
    # Avoid literal CURRENT_TIMESTAMP through dynamic insert if column defaults exist.
    columns = get_columns(cur, "recipes")
    if "created_at" in columns or "updated_at" in columns:
        filtered = {k: v for k, v in data.items() if k in columns and k not in {"created_at", "updated_at"}}
        col_sql = ", ".join(list(filtered.keys()) + [c for c in ["created_at", "updated_at"] if c in columns])
        placeholders = ", ".join(["?"] * len(filtered) + ["CURRENT_TIMESTAMP"] * len([c for c in ["created_at", "updated_at"] if c in columns]))
        cur.execute(f"INSERT INTO recipes ({col_sql}) VALUES ({placeholders})", tuple(filtered.values()))
        return int(cur.lastrowid)
    return insert_dynamic(cur, "recipes", data)


def create_recipe_parameter_values(cur: sqlite3.Cursor, recipe_id: int, parameter_ids_by_tag: dict[int, tuple[int, float]]) -> int:
    count = 0
    for _tag_index, (parameter_id, default_value) in sorted(parameter_ids_by_tag.items()):
        insert_dynamic(
            cur,
            "recipe_parameter_values",
            {
                "recipe_id": recipe_id,
                "parameter_id": parameter_id,
                "parameter_value": default_value,
            },
        )
        count += 1
    return count


def create_recipe_phase_rows(cur: sqlite3.Cursor, recipe_id: int, machine_stage_id: int, phase_rows: list[dict[str, Any]]) -> int:
    count = 0
    for row in phase_rows:
        phase_control_id = ensure_phase_master(
            cur,
            machine_stage_id,
            row["phase_group_code"],
            row["phase_group_name"],
            row["phase_control_name"],
            row.get("description") or row["phase_control_name"],
            int(row.get("display_order") or row["line_no"]),
        )
        insert_dynamic(
            cur,
            "recipe_phase_control",
            {
                "recipe_id": recipe_id,
                "line_no": int(row["line_no"]),
                "phase_control_id": phase_control_id,
                "phase_group_code": row["phase_group_code"],
                "phase_group_name": row["phase_group_name"],
                "sequence_no": int(row["line_no"]),
                "used": int(row.get("used") or 1),
            },
        )
        count += 1
    return count


def log_audit_if_possible(
    cur: sqlite3.Cursor,
    username: str,
    recipe_code: str,
    recipe_version: int,
    recipe_id: int,
    parameter_count: int,
    phase_count: int,
) -> None:
    if not table_exists(cur, "audit_log"):
        return
    data = {
        "username": username,
        "role": "ADMIN",
        "action": "P15_SECOND_STAGE_DRAFT_RECIPE_CREATED",
        "change_source": "P15_SECOND_STAGE_PARAMETER_MASTER_DRAFT_RECIPE_V1",
        "recipe_id": recipe_id,
        "recipe_code": recipe_code,
        "recipe_version": recipe_version,
        "old_value": None,
        "new_value": f"DRAFT created with {parameter_count} parameters and {phase_count} phase rows",
        "reason": "P15 Second Stage Parameter Master + Draft Recipe V1",
        "timestamp": "CURRENT_TIMESTAMP",
    }
    columns = get_columns(cur, "audit_log")
    filtered = {k: v for k, v in data.items() if k in columns and k != "timestamp"}
    timestamp_col = "timestamp" if "timestamp" in columns else None
    if timestamp_col:
        col_sql = ", ".join(list(filtered.keys()) + [timestamp_col])
        placeholders = ", ".join(["?"] * len(filtered) + ["CURRENT_TIMESTAMP"])
        cur.execute(f"INSERT INTO audit_log ({col_sql}) VALUES ({placeholders})", tuple(filtered.values()))
    else:
        insert_dynamic(cur, "audit_log", filtered)


def verify_created_recipe(cur: sqlite3.Cursor, recipe_id: int) -> dict[str, Any]:
    recipe = cur.execute(
        """
        SELECT r.id, r.recipe_code, r.recipe_name, r.version, r.status, r.machine_id, r.stage_id,
               m.machine_code, s.stage_type
        FROM recipes r
        LEFT JOIN tbm_machines m ON m.id=r.machine_id
        LEFT JOIN machine_stages s ON s.id=r.stage_id
        WHERE r.id=?
        """,
        (recipe_id,),
    ).fetchone()
    if not recipe:
        raise RuntimeError(f"Recipe not found after insert: {recipe_id}")
    parameter_count = cur.execute(
        "SELECT COUNT(*) c FROM recipe_parameter_values WHERE recipe_id=?",
        (recipe_id,),
    ).fetchone()["c"]
    phase_count = cur.execute(
        "SELECT COUNT(*) c FROM recipe_phase_control WHERE recipe_id=?",
        (recipe_id,),
    ).fetchone()["c"]
    phase_groups = cur.execute(
        """
        SELECT phase_group_code, COUNT(*) c
        FROM recipe_phase_control
        WHERE recipe_id=?
        GROUP BY phase_group_code
        ORDER BY phase_group_code
        """,
        (recipe_id,),
    ).fetchall()
    result = dict(recipe)
    result["parameter_count"] = int(parameter_count)
    result["phase_count"] = int(phase_count)
    result["phase_groups"] = [dict(row) for row in phase_groups]
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Build P15 SECOND_STAGE parameter master and draft recipe from Excel.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL_PATH), help="Input Excel path")
    parser.add_argument("--created-by", default=DEFAULT_CREATED_BY, help="Created by username")
    parser.add_argument("--recipe-code", default=None, help="Override recipe code")
    parser.add_argument("--recipe-name", default=None, help="Override recipe name")
    parser.add_argument("--version", type=int, default=None, help="Override recipe version")
    parser.add_argument("--dry-run", action="store_true", help="Validate only; do not write database")
    args = parser.parse_args()

    if not DB_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DB_PATH}")

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = PROJECT_ROOT / excel_path

    wb = read_workbook(excel_path)
    info = read_recipe_info(wb)
    recipe_code = args.recipe_code or info["recipe_code"]
    recipe_name = args.recipe_name or info["recipe_name"]
    recipe_version = args.version if args.version is not None else int(info["version"])
    created_by = args.created_by or info.get("created_by") or DEFAULT_CREATED_BY

    parameters = read_parameters(wb)
    excel_phase_rows = read_phase_rows(wb)

    print(f"Excel input       : {excel_path}")
    print(f"Recipe code      : {recipe_code}")
    print(f"Recipe name      : {recipe_name}")
    print(f"Recipe version   : {recipe_version}")
    print(f"Parameter rows   : {len(parameters)}")
    print(f"Phase rows input : {len(excel_phase_rows)}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        for table in [
            "tbm_machines",
            "machine_stages",
            "parameter_definitions",
            "recipes",
            "recipe_parameter_values",
            "recipe_phase_control",
            "phase_control_master",
        ]:
            if not table_exists(cur, table):
                raise RuntimeError(f"Required table missing: {table}")

        machine_id, machine_stage_id = resolve_p15_second_stage(cur)
        print(f"Resolved P15 machine_id      : {machine_id}")
        print(f"Resolved P15 SECOND_STAGE id : {machine_stage_id}")

        ensure_phase_schema(cur)
        ensure_phase_group_masters(cur, machine_stage_id)
        ensure_default_phase_masters(cur, machine_stage_id)

        if not excel_phase_rows:
            phase_rows = get_default_phase_rows_from_db(cur, machine_stage_id)
            print(f"Phase rows source: DB default phase masters ({len(phase_rows)} rows)")
        else:
            phase_rows = excel_phase_rows
            print(f"Phase rows source: Excel Phase_Control sheet ({len(phase_rows)} rows)")

        if not phase_rows:
            raise ValidationError("No phase-control rows available for P15 SECOND_STAGE.")

        duplicate = cur.execute(
            """
            SELECT id, status
            FROM recipes
            WHERE recipe_code=? AND version=? AND machine_id=? AND stage_id=?
            """,
            (recipe_code, recipe_version, machine_id, machine_stage_id),
        ).fetchone()
        if duplicate:
            raise ValidationError(
                f"Recipe already exists: id={duplicate['id']}, status={duplicate['status']}. "
                "Use --recipe-code or --version with a new value."
            )

        parameter_ids_by_tag: dict[int, tuple[int, float]] = {}
        for row in parameters:
            parameter_id = upsert_parameter_definition(cur, machine_id, machine_stage_id, row, created_by)
            parameter_ids_by_tag[int(row["tag_index"])] = (parameter_id, float(row["default_value"] or 0.0))

        if args.dry_run:
            conn.rollback()
            print("Dry run passed. No database changes committed.")
            return

        recipe_id = create_draft_recipe(
            cur,
            machine_id,
            machine_stage_id,
            recipe_code,
            recipe_name,
            recipe_version,
            created_by,
        )
        parameter_value_count = create_recipe_parameter_values(cur, recipe_id, parameter_ids_by_tag)
        phase_value_count = create_recipe_phase_rows(cur, recipe_id, machine_stage_id, phase_rows)
        log_audit_if_possible(
            cur,
            created_by,
            recipe_code,
            recipe_version,
            recipe_id,
            parameter_value_count,
            phase_value_count,
        )

        conn.commit()

        summary = verify_created_recipe(cur, recipe_id)
        print("\nCreated P15 SECOND_STAGE DRAFT recipe")
        print("-------------------------------------")
        print(summary)
        print("\nP15 Second Stage Parameter Master + Draft Recipe V1 completed successfully.")

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    try:
        main()
    except ValidationError as exc:
        print("\nVALIDATION FAILED")
        print("-----------------")
        print(exc)
        sys.exit(2)
